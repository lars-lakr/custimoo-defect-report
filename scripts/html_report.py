#!/usr/bin/env python3
"""Generate interactive HTML Defect Report with drill-down."""
import sys, os, json, pymysql
from collections import defaultdict
from decimal import Decimal

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
import factory_data

# Generate summary data
data = factory_data.generate()

months = data['months']
total_monthly = {m: {'qty': v['qty'], 'orders': v.get('orders', 0)} for m, v in data['total_monthly'].items()}
monthly_defects = data['monthly_defects']
defect_order_count = data.get('defect_order_count', {})
remake_by_month = data.get('remake_by_month', {})
factories = data['factories']
factory_monthly_data = data['factory_monthly']


def norm_qarma_supplier(name):
    n = (name or '').strip()
    low = n.lower()
    if 'mavic' in low:
        return 'Mavic Sports'
    if 'silver' in low:
        return 'Silver-Star Group'
    if 'selber' in low or 'seleber' in low:
        return 'Selberian Sports Wear'
    if 'karrizo' in low or 'karizzo' in low:
        return 'Karrizo'
    if 'rajco' in low:
        return 'Rajco'
    if 'custimoo' in low:
        return 'Custimoo factory'
    if 'augusta' in low:
        return 'Augusta De Mexico'
    return n or '(unknown)'

QARMA_STATS_CACHE = {}
QARMA_ORDER_STATS_CACHE = {}

def dt_to_month(v):
    import datetime
    if isinstance(v, datetime.datetime):
        return v.strftime('%Y-%m')
    if isinstance(v, datetime.date):
        return v.strftime('%Y-%m')
    if v:
        return str(v)[:7]
    return None

def load_qarma_stats(month_filter=None):
    """Aggregate Qarma physical QC by supplier for the report window.
    Sample qty is deduped by Report inspection id; defect pieces are summed from minor/major/critical pieces affected.
    """
    from pathlib import Path
    from collections import defaultdict
    cache_key = tuple(month_filter) if month_filter else tuple(months)
    if cache_key in QARMA_STATS_CACHE:
        return QARMA_STATS_CACHE[cache_key]
    try:
        import openpyxl
    except Exception:
        return {}
    export_dir = Path('/Users/lakr-macmini/Desktop/qarma')
    candidates = sorted(export_dir.glob('export*.xlsx'), key=lambda x: x.stat().st_mtime, reverse=True)
    path = candidates[0] if candidates else None
    if not path:
        return {}
    months_allowed = set(month_filter) if month_filter else set(months)
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    stats = defaultdict(lambda: {'sample_qty': 0, 'defects': 0, 'reports': set(), 'orders': set(), 'rejected_orders': set()})
    seen_report_sample = set()
    for r in ws.iter_rows(min_row=2, values_only=True):
        if len(r) < 38 or r[2] != 'Report' or str(r[9] or '').strip() != 'Final' or str(r[14] or '').strip() != 'Approved' or str(r[18] or '').strip().lower() == 'true' or not str(r[17] or '').strip().lower().endswith('@custimoo.com') or (len(r) >= 29 and r[28] is not None):
            continue
        month = dt_to_month(r[21])
        if month not in months_allowed:
            continue
        f = norm_qarma_supplier(r[8])
        report_id = str(r[1] or r[0] or '')
        order_no = str(r[3] or '')
        sample_qty = int(r[23] or 0)
        minor = int(r[35] or 0)
        major = int(r[36] or 0)
        critical = int(r[37] or 0)
        defects = minor + major + critical
        stats[f]['defects'] += defects
        if order_no and defects > 0:
            stats[f]['rejected_orders'].add(order_no)
        if order_no:
            stats[f]['orders'].add(order_no)
        if report_id:
            stats[f]['reports'].add(report_id)
            key = (f, report_id)
            if key not in seen_report_sample:
                seen_report_sample.add(key)
                stats[f]['sample_qty'] += sample_qty
    out = {}
    for f, v in stats.items():
        sample = v['sample_qty']
        defects = v['defects']
        out[f] = {
            'sample_qty': sample,
            'defects': defects,
            'rate': round(defects / sample * 100, 2) if sample > 0 else 0,
            'inspections': len(v['reports']),
            'orders_checked': len(v['orders']),
            'rejected_orders': len(v.get('rejected_orders', set())),
            'order_rate': round(len(v.get('rejected_orders', set())) / len(v['orders']) * 100, 2) if len(v['orders']) > 0 else 0,
        }
    QARMA_STATS_CACHE[cache_key] = out
    return out

def load_qarma_order_stats(month_filter=None):
    """Aggregate Qarma physical QC by backend order number for SKU/Admin groupings."""
    from pathlib import Path
    from collections import defaultdict
    cache_key = tuple(month_filter) if month_filter else tuple(months)
    if cache_key in QARMA_ORDER_STATS_CACHE:
        return QARMA_ORDER_STATS_CACHE[cache_key]
    try:
        import openpyxl
    except Exception:
        return {}
    export_dir = Path('/Users/lakr-macmini/Desktop/qarma')
    candidates = sorted(export_dir.glob('export*.xlsx'), key=lambda x: x.stat().st_mtime, reverse=True)
    path = candidates[0] if candidates else None
    if not path:
        return {}
    months_allowed = set(month_filter) if month_filter else set(months)
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    stats = defaultdict(lambda: {'sample_qty': 0, 'defects': 0, 'reports': set()})
    seen_report_sample = set()
    for r in ws.iter_rows(min_row=2, values_only=True):
        if len(r) < 38 or r[2] != 'Report' or str(r[9] or '').strip() != 'Final' or str(r[14] or '').strip() != 'Approved' or str(r[18] or '').strip().lower() == 'true' or not str(r[17] or '').strip().lower().endswith('@custimoo.com') or (len(r) >= 29 and r[28] is not None):
            continue
        month = dt_to_month(r[21])
        if month not in months_allowed:
            continue
        order_no = str(r[3] or '').strip()
        if not order_no:
            continue
        report_id = str(r[1] or r[0] or '')
        sample_qty = int(r[23] or 0)
        defects = int(r[35] or 0) + int(r[36] or 0) + int(r[37] or 0)
        stats[order_no]['defects'] += defects
        if report_id:
            stats[order_no]['reports'].add(report_id)
            key = (order_no, report_id)
            if key not in seen_report_sample:
                seen_report_sample.add(key)
                stats[order_no]['sample_qty'] += sample_qty
    out = {}
    for ono, v in stats.items():
        sample = v['sample_qty']
        defects = v['defects']
        out[ono] = {
            'sample_qty': sample,
            'defects': defects,
            'rate': round(defects / sample * 100, 2) if sample > 0 else 0,
            'orders_checked': 1,
            'rejected_orders': 1 if defects > 0 else 0,
            'order_rate': 100 if defects > 0 else 0,
            'inspections': len(v['reports']),
        }
    QARMA_ORDER_STATS_CACHE[cache_key] = out
    return out

qarma_stats = load_qarma_stats()

month_labels = {
    "2025-10": "Oct 2025", "2025-11": "Nov 2025", "2025-12": "Dec 2025",
    "2026-01": "Jan 2026", "2026-02": "Feb 2026", "2026-03": "Mar 2026",
    "2026-04": "Apr 2026", "2026-05": "May 2026", "2026-06": "Jun 2026*",
}

total_volume = sum(d['qty'] for d in total_monthly.values())
total_orders = sum(d['orders'] for d in total_monthly.values())
total_defects = sum(monthly_defects.values())
total_defect_orders = sum(defect_order_count.values())
total_rate = round(total_defects / total_volume * 100, 2) if total_volume > 0 else 0
total_order_rate = round(total_defect_orders / total_orders * 100, 2) if total_orders > 0 else 0

all_months_list = sorted(total_monthly.keys())
last_3 = all_months_list[-3:] if len(all_months_list) >= 3 else all_months_list
rolling_volume = sum(total_monthly.get(m, {}).get('qty', 0) for m in last_3)
rolling_orders = sum(total_monthly.get(m, {}).get('orders', 0) for m in last_3)
rolling_defects = sum(monthly_defects.get(m, 0) for m in last_3)
rolling_defect_orders = sum(defect_order_count.get(m, 0) for m in last_3)
rolling_rate = round(rolling_defects / rolling_volume * 100, 2) if rolling_volume > 0 else 0
rolling_order_rate = round(rolling_defect_orders / rolling_orders * 100, 2) if rolling_orders > 0 else 0
last_month_label = month_labels.get(last_3[-1], last_3[-1]) if last_3 else ""

all_months_sorted = sorted(total_monthly.keys())
report_month_labels = [month_labels.get(m, m) for m in all_months_sorted]

for f in factories:
    f['qarma'] = qarma_stats.get(f['name'], {'sample_qty': 0, 'defects': 0, 'rate': 0, 'inspections': 0, 'orders_checked': 0})

monthly_rows = []
for m in months:
    vol = total_monthly.get(m, {}).get('qty', 0)
    defs = monthly_defects.get(m, 0)
    rate = round(defs / vol * 100, 2) if vol > 0 else 0
    monthly_rows.append({'month': month_labels.get(m, m), 'volume': vol, 'defects': defs, 'rate': rate})

report_data = {
    'months': report_month_labels,
    'monthlyVolume': [r['volume'] for r in monthly_rows],
    'monthlyOrders': [total_monthly.get(m, {}).get('orders', 0) for m in months],
    'monthlyDefects': [r['defects'] for r in monthly_rows],
    'monthlyDefectOrders': [defect_order_count.get(m, 0) for m in months],
    'monthlyRemakeOrders': [remake_by_month.get(m, {}).get('orders', 0) for m in months],
    'monthlyRemakeQty': [remake_by_month.get(m, {}).get('qty', 0) for m in months],
    'monthlyRate': [r['rate'] for r in monthly_rows],
    'totalVolume': total_volume,
    'totalOrders': total_orders,
    'totalDefects': total_defects,
    'totalDefectOrders': total_defect_orders,
    'totalRate': total_rate,
    'totalOrderRate': total_order_rate,
    'rollingRate': rolling_rate,
    'rollingOrderRate': rolling_order_rate,
    'rollingLabel': (month_labels.get(last_3[0], last_3[0]) + " – " + month_labels.get(last_3[-1], last_3[-1])) if len(last_3) >= 2 else '',
    'factories': factories,
    'factoryMonthly': factory_monthly_data,
}

# ── YTD 2026 data (Jan–Jun) ──
ytd_months = [m for m in all_months_sorted if m.startswith("2026")]
ytd_month_labels = [month_labels.get(m, m) for m in ytd_months]
ytd_volume = sum(total_monthly.get(m, {}).get('qty', 0) for m in ytd_months)
ytd_orders = sum(total_monthly.get(m, {}).get('orders', 0) for m in ytd_months)
ytd_defects = sum(monthly_defects.get(m, 0) for m in ytd_months)
ytd_defect_orders = sum(defect_order_count.get(m, 0) for m in ytd_months)
ytd_rate = round(ytd_defects / ytd_volume * 100, 2) if ytd_volume > 0 else 0
ytd_order_rate = round(ytd_defect_orders / ytd_orders * 100, 2) if ytd_orders > 0 else 0

ytd_qarma_stats = load_qarma_stats(ytd_months)

# YTD factory totals
ytd_factories = []
for f in factories:
    fname = f['name']
    # Recalculate using only 2026 months
    fd = next((fd for fd in factory_monthly_data if fd['name'] == fname), None)
    if fd:
        ytd_def = sum(fd['defects'][i] for i, m in enumerate(all_months_sorted) if m.startswith("2026") and i < len(fd['defects']))
        ytd_vol = sum(fd['volumes'][i] for i, m in enumerate(all_months_sorted) if m.startswith("2026") and i < len(fd['volumes']))
        ytd_orders_f = sum(fd.get('orders', [])[i] for i, m in enumerate(all_months_sorted) if m.startswith("2026") and i < len(fd.get('orders', [])))
        ytd_defect_orders_f = sum(fd.get('defect_orders', [])[i] for i, m in enumerate(all_months_sorted) if m.startswith("2026") and i < len(fd.get('defect_orders', [])))
        ytd_remake_orders_f = sum(fd.get('remake_orders', [])[i] for i, m in enumerate(all_months_sorted) if m.startswith("2026") and i < len(fd.get('remake_orders', [])))
        ytd_remake_qty_f = sum(fd.get('remake_qty', [])[i] for i, m in enumerate(all_months_sorted) if m.startswith("2026") and i < len(fd.get('remake_qty', [])))
        ytd_rate_f = round(ytd_def / ytd_vol * 100, 2) if ytd_vol > 0 else 0
        ytd_order_rate_f = round(ytd_defect_orders_f / ytd_orders_f * 100, 2) if ytd_orders_f > 0 else 0
        ytd_factories.append({'name': fname, 'volume': ytd_vol, 'orders': ytd_orders_f, 'defects': ytd_def, 'defect_orders': ytd_defect_orders_f, 'remake_orders': ytd_remake_orders_f, 'remake_qty': ytd_remake_qty_f, 'rate': ytd_rate_f, 'order_rate': ytd_order_rate_f, 'qarma': ytd_qarma_stats.get(fname, {'sample_qty': 0, 'defects': 0, 'rate': 0, 'inspections': 0, 'orders_checked': 0})})
ytd_factories.sort(key=lambda x: -x['rate'])


DATA_JSON = json.dumps(report_data, cls=factory_data.DecimalEncoder)
YTD_DATA_JSON = json.dumps({
    'months': ytd_month_labels,
    'monthKeys': ytd_months,
    'monthlyVolume': [total_monthly.get(m, {}).get('qty', 0) for m in ytd_months],
    'monthlyOrders': [total_monthly.get(m, {}).get('orders', 0) for m in ytd_months],
    'monthlyDefects': [monthly_defects.get(m, 0) for m in ytd_months],
    'monthlyDefectOrders': [defect_order_count.get(m, 0) for m in ytd_months],
    'volume': ytd_volume,
    'orders': ytd_orders,
    'defects': ytd_defects,
    'defectOrders': ytd_defect_orders,
    'rate': ytd_rate,
    'orderRate': ytd_order_rate,
    'cumulativeVolume': [sum(total_monthly.get(m2, {}).get('qty', 0) for m2 in ytd_months[:i+1]) for i in range(len(ytd_months))],
    'cumulativeOrders': [sum(total_monthly.get(m2, {}).get('orders', 0) for m2 in ytd_months[:i+1]) for i in range(len(ytd_months))],
    'cumulativeDefects': [sum(monthly_defects.get(m2, 0) for m2 in ytd_months[:i+1]) for i in range(len(ytd_months))],
    'cumulativeDefectOrders': [sum(defect_order_count.get(m2, 0) for m2 in ytd_months[:i+1]) for i in range(len(ytd_months))],
    'cumulativeRate': [round(
        sum(monthly_defects.get(m2, 0) for m2 in ytd_months[:i+1]) /
        sum(total_monthly.get(m2, {}).get('qty', 0) for m2 in ytd_months[:i+1]) * 100, 2
    ) if sum(total_monthly.get(m2, {}).get('qty', 0) for m2 in ytd_months[:i+1]) > 0 else 0 for i in range(len(ytd_months))],
    'cumulativeOrderRate': [round(
        sum(defect_order_count.get(m2, 0) for m2 in ytd_months[:i+1]) /
        sum(total_monthly.get(m2, {}).get('orders', 0) for m2 in ytd_months[:i+1]) * 100, 2
    ) if sum(total_monthly.get(m2, {}).get('orders', 0) for m2 in ytd_months[:i+1]) > 0 else 0 for i in range(len(ytd_months))],
    'factories': ytd_factories,
}, cls=factory_data.DecimalEncoder)

FACTORY_COLORS = json.dumps({
    "Mavic Sports": "rgba(217, 45, 32, 0.85)",
    "Selberian Sports Wear": "rgba(247, 144, 9, 0.85)",
    "Silver-Star Group": "rgba(124, 58, 237, 0.82)",
    "Karrizo": "rgba(18, 183, 106, 0.85)",
    "Rajco": "rgba(31, 111, 235, 0.85)",
})

# Load order-level data — use factory_data.MANUAL for consistency
import urllib.request, urllib.parse, re

pw = os.environ.get("CUSTIMOO_DB_PASSWORD", "")
conn = pymysql.connect(host="127.0.0.1", port=3307, database="custimoo_backend_prod", user="custimoo_backend_usr", password=pw, connect_timeout=10)
cur = conn.cursor()

# Fetch fu messages (skip if no Graph credentials)
if os.environ.get("CUSTIMOO_GRAPH_TENANT_ID"):
    tenant = os.environ["CUSTIMOO_GRAPH_TENANT_ID"]
    client_id = os.environ["CUSTIMOO_GRAPH_CLIENT_ID"]
    client_secret = os.environ["CUSTIMOO_GRAPH_CLIENT_SECRET"]
    url = "https://login.microsoftonline.com/%s/oauth2/v2.0/token" % (tenant,)
    body = "client_id=%s&client_secret=%s&scope=https://graph.microsoft.com/.default&grant_type=client_credentials" % (client_id, client_secret)
    req = urllib.request.Request(url, data=body.encode(), headers={"Content-Type": "application/x-www-form-urlencoded"})
    try:
        token = json.loads(urllib.request.urlopen(req).read().decode())["access_token"]

        url = "https://graph.microsoft.com/v1.0/users/fu@custimoo.com/messages?$filter=receivedDateTime%20ge%202025-10-28T00:00:00Z&$top=100&$select=subject,receivedDateTime,body&$orderby=receivedDateTime%20asc"
        all_msgs = []
        while url:
            req = urllib.request.Request(url, headers={"Authorization": "Bearer " + token, "Prefer": "outlook.body-content-type=text"})
            resp = json.loads(urllib.request.urlopen(req).read().decode())
            all_msgs.extend(resp.get("value", []))
            url = resp.get("@odata.nextLink")

        first_fu_month = {}
        groups = defaultdict(list)
        for m in all_msgs:
            subj = m["subject"]
            content = m.get("body", {}).get("content", "")[:2000]
            dt = m["receivedDateTime"][:7]
            nums_text = re.sub(r"https?://\S+", " ", subj + " " + content)
            nums = re.findall(r"(?:order|Order|ORDER)?\s*#?\s*(\d{4,6})", nums_text)
            for n in set(n for n in nums if 10000 <= int(n) <= 99999):
                groups[n].append({"subject": subj, "body": content})
                if n not in first_fu_month:
                    first_fu_month[n] = dt
    except Exception as e:
        print("Graph API call in html_report.py failed:", e)
        first_fu_month = {}
        groups = defaultdict(list)
else:
    print("Skipping Microsoft Graph in html_report — no credentials")
    first_fu_month = {}
    groups = defaultdict(list)

ALL_ORDER_NUMS = set(list(first_fu_month.keys()) + list(factory_data.MANUAL.keys()))
qs = ",".join(["%s"] * len(ALL_ORDER_NUMS))

cur.execute("SELECT o.order_no, CAST(JSON_EXTRACT(o.price_info, '$.total_quantity') AS SIGNED), COALESCE(oi.factory_name, '(unknown)') FROM orders o LEFT JOIN order_items oi ON oi.order_id = o.id WHERE o.order_no IN (%s)" % qs, list(ALL_ORDER_NUMS))

order_factory = {}
order_qty = {}
db_order_nums = set()
for r in cur.fetchall():
    ono = str(r[0])
    db_order_nums.add(ono)
    order_factory[ono] = factory_data.norm_factory(r[2])
    order_qty[ono] = r[1] or 0

from report import classify_product, extract_qty
product_types = {}
for ono in ALL_ORDER_NUMS:
    cat, is_prod = classify_product(ono, conn)
    product_types[ono] = cat

def categorize_root_cause(root, issue='', corrective='', remarks=''):
    """Return (category, confidence). Only >=90 is accepted; otherwise Uncategorized."""
    import re
    text = ' '.join(str(x or '') for x in (root, issue, corrective, remarks)).lower()
    text = re.sub(r'\s+', ' ', text)
    if not text.strip():
        return 'Uncategorized', 0
    rules = [
        ('Missing item / packing', 96, ['missing pcs', 'missing jerseys', 'missing pieces', 'additional cemara', 'packing table', 'camera added into packing', 'camara added into packing']),
        ('Missing branding / logo', 95, ['missing branding', 'brand logo', 'branding logo was missed', 'branding logos added', 'logo was missed']),
        ('Wrong badge / logo', 94, ['wrong badge', 'badge color incorrect', 'badge colour incorrect', 'produced badge', 'prduced badge', 'wrong logo', 'logo size', '3d logo peel', 'wrong colour in logo', 'wrong color in logo']),
        ('Wrong number / sizing', 94, ['number sizing wrong', 'wrong number', 'wrong size', 'size of logo', 'sizes wrong']),
        ('Spec / tech pack mismatch', 93, ['tds was correct', 'teck pack was wrong', 'tech pack was wrong', 'only check tds', 'as per teckpack', 'as per techpack', 'customer not informed', 'not aware', 'rawling specs']),
        ('Embroidery / decoration method', 94, ['embroidery', 'emb.', '3000/4000', 'sublimation', 'twill patch', 'zigzag']),
        ('Color issue', 93, ['wrong color', 'wrong colour', 'buttons color', 'pantone', 'color incorrect', 'colour incorrect']),
        ('Garment construction / stitching', 94, ['broken stitch', 'stitch at zips', 'defective zippers', 'notching', 'sleeve', 'zipper', 'zips']),
        ('Fabric / material', 94, ['fabric', 'felt cheaper', 'approve sample fabric']),
        ('Accessory / trims', 93, ['necktape', 'neck tape', 'tag', 'buttons']),
        ('Factory supplied component', 92, ['provided woven label', 'sourcing anything', 'provided 3d logos', 'we provided']),
        ('No production issue', 95, ['no issue with production', 'was correct', 'shipped / video provided', 'not shipped yet']),
        ('QC overlooked', 91, ['qc overlooked', 'overlooked by qc', 'qc guy', 'qc for more focus', 'during final inspection', 'aql audit']),
    ]
    for category, confidence, needles in rules:
        if any(n in text for n in needles):
            return category, confidence
    return 'Uncategorized', 0

def load_issue_categories():
    import re
    from pathlib import Path
    import openpyxl
    candidates = sorted(Path('/Users/lakr-macmini/Desktop/qarma').glob('Order*Issues*Monthly*.xlsx'), key=lambda x: x.stat().st_mtime, reverse=True)
    if not candidates:
        return {}
    ws = openpyxl.load_workbook(candidates[0], read_only=True, data_only=True).active
    out = {}
    for r in ws.iter_rows(min_row=2, values_only=True):
        month = str(r[0] or '').strip()
        order = str(r[1] or '').strip()
        if not month or not order:
            continue
        ml = month.lower()
        if 'subtotal' in ml or 'grand total' in ml or '—' in month:
            continue
        m = re.search(r'\d{4,6}', order)
        if not m:
            continue
        order_no = m.group(0)
        root = r[8] if len(r) > 8 else None
        issue = r[6] if len(r) > 6 else None
        corrective = r[12] if len(r) > 12 else None
        remarks = r[16] if len(r) > 16 else None
        category, confidence = categorize_root_cause(root, issue, corrective, remarks)
        if confidence < 90:
            category = 'Uncategorized'
        out[order_no] = {
            'month': month,
            'issue': issue or '',
            'root_cause': root or '',
            'corrective_action': corrective or '',
            'remarks': remarks or '',
            'category': category,
            'confidence': confidence,
        }
    return out

ISSUE_CATEGORIES = load_issue_categories()



order_details = []
for ono in ALL_ORDER_NUMS:
    related_msgs = groups.get(ono, [])
    if ono in factory_data.MANUAL:
        affected = factory_data.MANUAL[ono]
        if affected == 0:
            continue
        source = 'manual'
    else:
        if factory_data.is_non_defect_fu(related_msgs):
            continue
        affected = extract_qty(related_msgs)
        if affected is None:
            continue
        source = 'parsed_fu'

    fu_month = first_fu_month.get(ono, "?")
    factory = order_factory.get(ono, "(unknown)")
    if ono in factory_data.MANUAL_FACTORY:
        factory = factory_data.MANUAL_FACTORY[ono]
    if factory in getattr(factory_data, 'EXCLUDED_FACTORIES', set()): continue
    if factory == "(unknown)": continue
    if ono not in db_order_nums: continue  # phantom ID from email URL, not a real order
    qty = order_qty.get(ono, 0)
    if source == 'parsed_fu' and qty and affected > qty:
        continue
    pt = product_types.get(ono, "Unknown")
    
    subjects = " | ".join(set(m["subject"] for m in related_msgs))[:200]
    snippet = (related_msgs[0]["body"][:300] if related_msgs else "")[:300]

    order_details.append({
        "order": ono,
        "affected": affected,
        "total_qty": qty,
        "fu_month": fu_month,
        "factory": factory,
        "product_type": pt,
        "subjects": subjects,
        "snippet": snippet,
        "category": ISSUE_CATEGORIES.get(ono, {}).get("category", "Uncategorized"),
        "root_cause": ISSUE_CATEGORIES.get(ono, {}).get("root_cause", ""),
        "category_confidence": ISSUE_CATEGORIES.get(ono, {}).get("confidence", 0),
        "source": source,
    })

order_details.sort(key=lambda x: -x["affected"])

counted_orders = {d['order'] for d in order_details}
fu_review_rows = []
for ono in sorted(db_order_nums, key=lambda n: (first_fu_month.get(n, ''), n), reverse=True):
    related_msgs = groups.get(ono, [])
    if not related_msgs:
        continue
    if ono in counted_orders:
        status = 'Counted in FU defects'
        affected_review = next((d['affected'] for d in order_details if d['order'] == ono), '')
    elif ono in factory_data.MANUAL_ZERO:
        status = 'Excluded — manual non-defect'
        affected_review = ''
    elif factory_data.is_non_defect_fu(related_msgs):
        status = 'Excluded — delay/process, not physical defect'
        affected_review = ''
    else:
        parsed_qty = extract_qty(related_msgs)
        if parsed_qty is None:
            status = 'Needs affected qty review'
            affected_review = ''
        else:
            status = 'Parsed but not counted — check product/factory filter'
            affected_review = parsed_qty
    fu_review_rows.append({
        'order': ono,
        'fu_month': first_fu_month.get(ono, '?'),
        'status': status,
        'affected': affected_review,
        'factory': order_factory.get(ono, '(unknown)'),
        'total_qty': order_qty.get(ono, 0),
        'subjects': ' | '.join(set(m['subject'] for m in related_msgs))[:200],
        'snippet': (related_msgs[0]['body'][:300] if related_msgs else '')[:300],
    })

def collect_sku_text(obj):
    parts = []
    keys = {'sku_id','sku_name','sku_number','description','product_name','product_display_name','design_nick_name','new_product_name','style_name'}
    def walk(x):
        if isinstance(x, dict):
            for k, v in x.items():
                if k in keys and v is not None:
                    parts.append(str(v))
                elif isinstance(v, (dict, list)):
                    walk(v)
        elif isinstance(x, list):
            for v in x:
                walk(v)
    walk(obj)
    return ' | '.join(parts)

def clean_sku_text(t):
    import re, html
    t = html.unescape(t or '')
    t = re.sub(r'<[^>]+>', ' ', t)
    t = re.sub(r'\s+', ' ', t).strip()
    return t

def classify_sku_text(t):
    import re
    low = (t or '').lower()
    series = sorted(set(re.findall(r'\b([1234])000\s*(?:-|\s)?(?:series|se)?\b', low)))
    series = [x + '000' for x in series]
    sublimated = bool(re.search(r'\bsublimat(?:ed|ion|e)?\b|\bfully\s+sub\b|\bsub\b', low))
    return series, sublimated


# ── Summary breakdown groupings (SKU / Order Admin) ──
def classify_order_series_from_text(text):
    series, sublimated = classify_sku_text(text)
    return series or ['No series found'], sublimated

def rollup_sku_groups(series):
    out = set()
    for ser in series:
        if ser in ('1000', '2000'):
            out.add('Sublimation')
        elif ser in ('3000', '4000'):
            out.add('Embroidery')
        else:
            out.add(ser)
    return sorted(out)


def classify_sport_text(text):
    t = ' ' + (text or '').lower() + ' '
    sport_patterns = [
        ('Hockey', ['hockey']),
        ('Basketball', ['basketball']),
        ('Baseball', ['baseball', 'softball']),
        ('Football', ['football', 'gridiron']),
        ('Soccer', ['soccer', 'football jersey']),
        ('Lacrosse', ['lacrosse']),
        ('Volleyball', ['volleyball']),
        ('Rugby', ['rugby']),
        ('Cycling', ['cycling', 'biking', 'bike jersey']),
        ('Running', ['running', 'track and field', 'athletics']),
        ('Esports', ['esport', 'e-sport', 'gaming jersey']),
        ('Training', ['training', 'warmup', 'warm-up']),
    ]
    found = []
    for label, needles in sport_patterns:
        if any(n in t for n in needles):
            found.append(label)
    return sorted(set(found)) or ['No sport found']


# Build per-order SKU text and admin name for all backend orders in report window.
cur.execute("""
SELECT o.order_no, CAST(JSON_EXTRACT(o.price_info, '$.total_quantity') AS SIGNED) AS qty,
       oi.status_updated_at AS shipping_date,
       COALESCE(u.name, u.email, '(unknown)') AS admin_name,
       COALESCE(oi.factory_name, '(unknown)') AS raw_factory,
       oi.factory_products, oi.order_line
FROM orders o
LEFT JOIN users u ON u.id = o.order_administrator_id
LEFT JOIN order_items oi ON oi.order_id = o.id
WHERE oi.status_updated_at >= '2025-10-01'
  AND oi.status_updated_at < '2026-07-01'
  AND (oi.status IN ('shipped','completed') OR oi.shipping_status IS NOT NULL)
""")
all_order_meta = defaultdict(lambda: {'qty': 0, 'admin': '(unknown)', 'texts': [], 'month': '?'})
for ono, qty, shipping_date, admin_name, raw_factory, factory_products, order_line in cur.fetchall():
    if factory_data.norm_factory(raw_factory) in getattr(factory_data, 'EXCLUDED_FACTORIES', set()):
        continue
    ono = str(ono)
    all_order_meta[ono]['qty'] = max(all_order_meta[ono]['qty'], int(qty or 0))
    all_order_meta[ono]['admin'] = admin_name or '(unknown)'
    all_order_meta[ono]['month'] = str(shipping_date)[:7] if shipping_date else '?'
    for raw in (factory_products, order_line):
        if not raw:
            continue
        try:
            parsed = json.loads(raw)
        except Exception:
            continue
        txt = clean_sku_text(collect_sku_text(parsed))
        if txt:
            all_order_meta[ono]['texts'].append(txt)

def empty_group():
    return {'volume': 0, 'orders_set': set(), 'defects': 0, 'defect_orders_set': set(), 'remake_orders_set': set(), 'remake_qty_total': 0, 'qarma_order_set': set(), 'qarma': {'sample_qty': 0, 'defects': 0, 'rate': 0, 'orders_checked': 0, 'inspections': 0, 'rejected_orders': 0, 'order_rate': 0}}

def finalize_groups(groups):
    rows = []
    for name, g in groups.items():
        vol = g['volume']
        orders_count = len(g['orders_set'])
        defect_orders_count = len(g['defect_orders_set'])
        remake_orders_count = len(g['remake_orders_set'])
        remake_orders_qty = g.get('remake_qty_total', 0)
        rows.append({
            'name': name,
            'volume': vol,
            'orders': orders_count,
            'defects': g['defects'],
            'defect_orders': defect_orders_count,
            'remake_orders': remake_orders_count,
            'remake_qty': remake_orders_qty,
            'rate': round(g['defects'] / vol * 100, 2) if vol > 0 else 0,
            'order_rate': round(defect_orders_count / orders_count * 100, 2) if orders_count > 0 else 0,
            'qarma': {**g.get('qarma', {'sample_qty': 0, 'defects': 0, 'rate': 0, 'orders_checked': 0, 'inspections': 0, 'rejected_orders': 0, 'order_rate': 0}), 'orders_checked': len(g.get('qarma_order_set', set())), 'rate': round(g.get('qarma', {}).get('defects', 0) / g.get('qarma', {}).get('sample_qty', 0) * 100, 2) if g.get('qarma', {}).get('sample_qty', 0) > 0 else 0, 'order_rate': round(g.get('qarma', {}).get('rejected_orders', 0) / len(g.get('qarma_order_set', set())) * 100, 2) if len(g.get('qarma_order_set', set())) > 0 else 0},
        })
    rows.sort(key=lambda x: -x['rate'])
    return rows

# Build remake orders lookup
conn = __import__('pymysql').connect(host="127.0.0.1", port=3307, database="custimoo_backend_prod", user="custimoo_backend_usr", password=__import__('os').environ.get("CUSTIMOO_DB_PASSWORD", ""), connect_timeout=10)
remake_cur = conn.cursor()
remake_cur.execute("SELECT o.order_no FROM orders o WHERE o.order_type_symbol = 'R' AND o.created_at >= '2025-10-01' AND o.created_at < '2026-07-01'")
REMAKE_ORDERS = set(str(r[0]) for r in remake_cur.fetchall())
remake_cur.close()
# Don't close main conn — used later

sku_groups = defaultdict(empty_group)
sport_groups = defaultdict(empty_group)
admin_groups = defaultdict(empty_group)
category_groups = defaultdict(empty_group)
for ono, meta in all_order_meta.items():
    text = ' | '.join(meta.get('texts', []))
    series, _sublimated = classify_order_series_from_text(text)
    for ser in rollup_sku_groups(series):
        sku_groups[ser]['volume'] += meta['qty']
        sku_groups[ser]['orders_set'].add(ono)
    for sport in classify_sport_text(text):
        sport_groups[sport]['volume'] += meta['qty']
        sport_groups[sport]['orders_set'].add(ono)
    admin = meta.get('admin') or '(unknown)'
    admin_groups[admin]['volume'] += meta['qty']
    admin_groups[admin]['orders_set'].add(ono)
    is_remake = ono in REMAKE_ORDERS
    if is_remake:
        for ser in rollup_sku_groups(series):
            sku_groups[ser]['remake_orders_set'].add(ono)
        for sport in classify_sport_text(text):
            sport_groups[sport]['remake_orders_set'].add(ono)
        admin_groups[admin]['remake_orders_set'].add(ono)

order_lookup = {d['order']: d for d in order_details}
for ono, d in order_lookup.items():
    meta = all_order_meta.get(ono, {})
    text = ' | '.join(meta.get('texts', []))
    series, _sublimated = classify_order_series_from_text(text)
    for ser in rollup_sku_groups(series):
        sku_groups[ser]['defects'] += d.get('affected', 0)
        sku_groups[ser]['defect_orders_set'].add(ono)
    for sport in classify_sport_text(text):
        sport_groups[sport]['defects'] += d.get('affected', 0)
        sport_groups[sport]['defect_orders_set'].add(ono)
    admin = meta.get('admin') or '(unknown)'
    admin_groups[admin]['defects'] += d.get('affected', 0)
    admin_groups[admin]['defect_orders_set'].add(ono)
    cat = d.get('category') or ISSUE_CATEGORIES.get(ono, {}).get('category', 'Uncategorized')
    if cat != 'Uncategorized' or ono in ISSUE_CATEGORIES:
        category_groups[cat]['volume'] += meta.get('qty', 0)
        category_groups[cat]['orders_set'].add(ono)
        category_groups[cat]['defects'] += d.get('affected', 0)
        category_groups[cat]['defect_orders_set'].add(ono)

# Add Qarma physical QC metrics into the same SKU/Admin groups, by Qarma order number.
qarma_order_stats = load_qarma_order_stats()
for ono, q in qarma_order_stats.items():
    meta = all_order_meta.get(ono)
    if not meta:
        continue
    text = ' | '.join(meta.get('texts', []))
    series, _sublimated = classify_order_series_from_text(text)
    admin = meta.get('admin') or '(unknown)'
    for ser in series:
        sku_groups[ser]['qarma']['sample_qty'] += q.get('sample_qty', 0)
        sku_groups[ser]['qarma']['defects'] += q.get('defects', 0)
        sku_groups[ser]['qarma']['rejected_orders'] += q.get('rejected_orders', 0)
        sku_groups[ser]['qarma']['inspections'] += q.get('inspections', 0)
        sku_groups[ser]['qarma_order_set'].add(ono)
    admin_groups[admin]['qarma']['sample_qty'] += q.get('sample_qty', 0)
    admin_groups[admin]['qarma']['defects'] += q.get('defects', 0)
    admin_groups[admin]['qarma']['rejected_orders'] += q.get('rejected_orders', 0)
    admin_groups[admin]['qarma']['inspections'] += q.get('inspections', 0)
    admin_groups[admin]['qarma_order_set'].add(ono)
    cat = ISSUE_CATEGORIES.get(ono, {}).get('category')
    if cat:
        category_groups[cat]['qarma']['sample_qty'] += q.get('sample_qty', 0)
        category_groups[cat]['qarma']['defects'] += q.get('defects', 0)
        category_groups[cat]['qarma']['rejected_orders'] += q.get('rejected_orders', 0)
        category_groups[cat]['qarma']['inspections'] += q.get('inspections', 0)
        category_groups[cat]['qarma_order_set'].add(ono)

GROUPING_JSON = json.dumps({
    'sku': finalize_groups(sku_groups),
    'sport': finalize_groups(sport_groups),
    'admin': finalize_groups(admin_groups),
    'category': finalize_groups(category_groups),
}, default=str)
GROUPING_JSON_SAFE = GROUPING_JSON.replace('<', '\\u003C').replace('>', '\\u003E')

def build_error_tracking():
    from collections import defaultdict
    et = {
        'factory': defaultdict(lambda: defaultdict(lambda: {'order_count': 0, 'defect_qty': 0, 'order_nums': []})),
        'sku': defaultdict(lambda: defaultdict(lambda: {'order_count': 0, 'defect_qty': 0, 'order_nums': []})),
        'sport': defaultdict(lambda: defaultdict(lambda: {'order_count': 0, 'defect_qty': 0, 'order_nums': []})),
        'admin': defaultdict(lambda: defaultdict(lambda: {'order_count': 0, 'defect_qty': 0, 'order_nums': []})),
    }
    for d in order_details:
        ono = d['order']
        cat_info = ISSUE_CATEGORIES.get(ono)
        if not cat_info:
            continue
        category = cat_info['category']
        defect = d.get('affected', 0)
        meta = all_order_meta.get(ono, {})
        text = ' | '.join(meta.get('texts', []))
        factory = d.get('factory', 'Unknown')
        et['factory'][factory][category]['order_count'] += 1
        et['factory'][factory][category]['defect_qty'] += defect
        et['factory'][factory][category]['order_nums'].append(ono)
        admin = meta.get('admin', 'Unknown')
        et['admin'][admin][category]['order_count'] += 1
        et['admin'][admin][category]['defect_qty'] += defect
        et['admin'][admin][category]['order_nums'].append(ono)
        series, _ = classify_order_series_from_text(text)
        for grp in rollup_sku_groups(series):
            et['sku'][grp][category]['order_count'] += 1
            et['sku'][grp][category]['defect_qty'] += defect
            et['sku'][grp][category]['order_nums'].append(ono)
        for sport in classify_sport_text(text):
            et['sport'][sport][category]['order_count'] += 1
            et['sport'][sport][category]['defect_qty'] += defect
            et['sport'][sport][category]['order_nums'].append(ono)
    result = {}
    for mode in ('factory','sku','sport','admin'):
        result[mode] = {}
        for group in sorted(et[mode]):
            cats = sorted(et[mode][group].items(), key=lambda x: -x[1]['defect_qty'])
            result[mode][group] = [{'category':c,'order_count':info['order_count'],'defect_qty':info['defect_qty'],'order_nums':info['order_nums']} for c,info in cats]
    return result

ERROR_TRACKING = build_error_tracking()
ERROR_TRACKING_JSON = json.dumps(ERROR_TRACKING, default=str)
ERROR_TRACKING_JSON_SAFE = ERROR_TRACKING_JSON.replace('<','\\u003C').replace('>','\\u003E')


# ── Summary period slices ──
def period_months_for(key):
    ordered = all_months_sorted
    current = ordered[-1] if ordered else ''
    prev = ordered[-2] if len(ordered) >= 2 else current
    if key == 'all':
        return ordered
    if key == 'last_3':
        return ordered[-3:]
    if key == 'last_6':
        return ordered[-6:]
    if key == 'last_month':
        return [prev]
    if key == 'mtd':
        return [current]
    if key == 'ytd':
        year = current[:4]
        return [m for m in ordered if m.startswith(year)]
    if key == 'quarter':
        year = current[:4]
        month_num = int(current[5:7]) if current and current[5:7].isdigit() else 1
        q_start = ((month_num - 1) // 3) * 3 + 1
        wanted = {f"{year}-{m:02d}" for m in range(q_start, q_start + 3)}
        return [m for m in ordered if m in wanted]
    return ordered

def period_label(month_keys):
    if not month_keys:
        return ''
    if len(month_keys) == 1:
        return month_labels.get(month_keys[0], month_keys[0])
    return month_labels.get(month_keys[0], month_keys[0]) + ' – ' + month_labels.get(month_keys[-1], month_keys[-1])

def qarma_empty():
    return {'sample_qty': 0, 'defects': 0, 'rate': 0, 'inspections': 0, 'orders_checked': 0, 'rejected_orders': 0, 'order_rate': 0}

def factory_rows_for_months(month_keys):
    qstats = load_qarma_stats(month_keys)
    rows = []
    month_set = set(month_keys)
    for fd in factory_monthly_data:
        vols = [fd.get('volumes', [])[i] for i, m in enumerate(all_months_sorted) if m in month_set and i < len(fd.get('volumes', []))]
        orders = [fd.get('orders', [])[i] for i, m in enumerate(all_months_sorted) if m in month_set and i < len(fd.get('orders', []))]
        defs = [fd.get('defects', [])[i] for i, m in enumerate(all_months_sorted) if m in month_set and i < len(fd.get('defects', []))]
        def_orders = [fd.get('defect_orders', [])[i] for i, m in enumerate(all_months_sorted) if m in month_set and i < len(fd.get('defect_orders', []))]
        remake_orders = [fd.get('remake_orders', [])[i] for i, m in enumerate(all_months_sorted) if m in month_set and i < len(fd.get('remake_orders', []))]
        remake_qty = [fd.get('remake_qty', [])[i] for i, m in enumerate(all_months_sorted) if m in month_set and i < len(fd.get('remake_qty', []))]
        vol = sum(vols); ords = sum(orders); defect = sum(defs); defect_orders = sum(def_orders)
        remake_orders_total = sum(remake_orders); remake_qty_total = sum(remake_qty)
        if vol == 0 and defect == 0:
            continue
        rows.append({
            'name': fd['name'], 'volume': vol, 'orders': ords, 'defects': defect,
            'defect_orders': defect_orders,
            'remake_orders': remake_orders_total,
            'remake_qty': remake_qty_total,
            'rate': round(defect / vol * 100, 2) if vol > 0 else 0,
            'order_rate': round(defect_orders / ords * 100, 2) if ords > 0 else 0,
            'qarma': qstats.get(fd['name'], qarma_empty()),
            'monthly': {'volumes': vols, 'defects': defs, 'orders': orders, 'defect_orders': def_orders}
        })
    rows.sort(key=lambda x: -x['rate'])
    return rows

def build_groupings_for_months(month_keys):
    month_set = set(month_keys)
    sku_groups = defaultdict(empty_group)
    sport_groups = defaultdict(empty_group)
    admin_groups = defaultdict(empty_group)
    category_groups = defaultdict(empty_group)
    for ono, meta in all_order_meta.items():
        if meta.get('month') not in month_set:
            continue
        text = ' | '.join(meta.get('texts', []))
        is_remake = ono in REMAKE_ORDERS
        series, _sublimated = classify_order_series_from_text(text)
        for ser in rollup_sku_groups(series):
            sku_groups[ser]['volume'] += meta['qty']
            sku_groups[ser]['orders_set'].add(ono)
            if is_remake:
                sku_groups[ser]['remake_orders_set'].add(ono)
                sku_groups[ser]['remake_qty_total'] += meta['qty']
        for sport in classify_sport_text(text):
            sport_groups[sport]['volume'] += meta['qty']
            sport_groups[sport]['orders_set'].add(ono)
            if is_remake:
                sport_groups[sport]['remake_orders_set'].add(ono)
                sport_groups[sport]['remake_qty_total'] += meta['qty']
        admin = meta.get('admin') or '(unknown)'
        admin_groups[admin]['volume'] += meta['qty']
        admin_groups[admin]['orders_set'].add(ono)
        if is_remake:
            admin_groups[admin]['remake_orders_set'].add(ono)
            admin_groups[admin]['remake_qty_total'] += meta['qty']
    for ono, d in order_lookup.items():
        if d.get('fu_month') not in month_set:
            continue
        meta = all_order_meta.get(ono, {})
        text = ' | '.join(meta.get('texts', []))
        series, _sublimated = classify_order_series_from_text(text)
        for ser in rollup_sku_groups(series):
            sku_groups[ser]['defects'] += d.get('affected', 0)
            sku_groups[ser]['defect_orders_set'].add(ono)
        for sport in classify_sport_text(text):
            sport_groups[sport]['defects'] += d.get('affected', 0)
            sport_groups[sport]['defect_orders_set'].add(ono)
        admin = meta.get('admin') or '(unknown)'
        admin_groups[admin]['defects'] += d.get('affected', 0)
        admin_groups[admin]['defect_orders_set'].add(ono)
        cat = d.get('category') or ISSUE_CATEGORIES.get(ono, {}).get('category', 'Uncategorized')
        if cat != 'Uncategorized' or ono in ISSUE_CATEGORIES:
            category_groups[cat]['volume'] += meta.get('qty', 0)
            category_groups[cat]['orders_set'].add(ono)
            category_groups[cat]['defects'] += d.get('affected', 0)
            category_groups[cat]['defect_orders_set'].add(ono)
    qstats_order = load_qarma_order_stats(month_keys)
    for ono, q in qstats_order.items():
        meta = all_order_meta.get(ono)
        if not meta or meta.get('month') not in month_set:
            continue
        text = ' | '.join(meta.get('texts', []))
        series, _sublimated = classify_order_series_from_text(text)
        admin = meta.get('admin') or '(unknown)'
        for ser in rollup_sku_groups(series):
            sku_groups[ser]['qarma']['sample_qty'] += q.get('sample_qty', 0)
            sku_groups[ser]['qarma']['defects'] += q.get('defects', 0)
            sku_groups[ser]['qarma']['rejected_orders'] += q.get('rejected_orders', 0)
            sku_groups[ser]['qarma']['inspections'] += q.get('inspections', 0)
            sku_groups[ser]['qarma_order_set'].add(ono)
        for sport in classify_sport_text(text):
            sport_groups[sport]['qarma']['sample_qty'] += q.get('sample_qty', 0)
            sport_groups[sport]['qarma']['defects'] += q.get('defects', 0)
            sport_groups[sport]['qarma']['rejected_orders'] += q.get('rejected_orders', 0)
            sport_groups[sport]['qarma']['inspections'] += q.get('inspections', 0)
            sport_groups[sport]['qarma_order_set'].add(ono)
        admin_groups[admin]['qarma']['sample_qty'] += q.get('sample_qty', 0)
        admin_groups[admin]['qarma']['defects'] += q.get('defects', 0)
        admin_groups[admin]['qarma']['rejected_orders'] += q.get('rejected_orders', 0)
        admin_groups[admin]['qarma']['inspections'] += q.get('inspections', 0)
        admin_groups[admin]['qarma_order_set'].add(ono)
        cat = ISSUE_CATEGORIES.get(ono, {}).get('category')
        if cat:
            category_groups[cat]['qarma']['sample_qty'] += q.get('sample_qty', 0)
            category_groups[cat]['qarma']['defects'] += q.get('defects', 0)
            category_groups[cat]['qarma']['rejected_orders'] += q.get('rejected_orders', 0)
            category_groups[cat]['qarma']['inspections'] += q.get('inspections', 0)
            category_groups[cat]['qarma_order_set'].add(ono)
    return {'sku': finalize_groups(sku_groups), 'sport': finalize_groups(sport_groups), 'admin': finalize_groups(admin_groups), 'category': finalize_groups(category_groups)}

def build_period_payload(key, display_name):
    mkeys = period_months_for(key)
    labels = [month_labels.get(m, m) for m in mkeys]
    vol = [total_monthly.get(m, {}).get('qty', 0) for m in mkeys]
    ords = [total_monthly.get(m, {}).get('orders', 0) for m in mkeys]
    defs = [monthly_defects.get(m, 0) for m in mkeys]
    def_orders = [defect_order_count.get(m, 0) for m in mkeys]
    rates = [round(defs[i] / vol[i] * 100, 2) if vol[i] > 0 else 0 for i in range(len(mkeys))]
    total_vol = sum(vol); total_orders_p = sum(ords); total_defs = sum(defs); total_def_orders = sum(def_orders)
    rows = factory_rows_for_months(mkeys)
    # Prev period
    prev = {}
    all_ord = all_months_sorted
    ci = {m:i for i,m in enumerate(all_ord)}
    n = len(mkeys)
    if n > 1 and n in (3,6):
        start_m = mkeys[0]
        si = ci.get(start_m, -1)
        if si >= n:
            pm = all_ord[si-n:si]
            pv = [total_monthly.get(m,{}).get('qty',0) for m in pm]
            po = [total_monthly.get(m,{}).get('orders',0) for m in pm]
            pd2 = [monthly_defects.get(m,0) for m in pm]
            pdo = [defect_order_count.get(m,0) for m in pm]
            pr = [round(pd2[i]/pv[i]*100,2) if pv[i]>0 else 0 for i in range(len(pm))]
            total_v = sum(pv); total_o = sum(po); total_d = sum(pd2); total_do = sum(pdo)
            prev = {
                'months': [month_labels.get(m,m) for m in pm],
                'monthKeys': pm,
                'monthlyVolume': pv, 'monthlyRate': pr,
                'monthlyDefects': pd2, 'monthlyOrders': po, 'monthlyDefectOrders': pdo,
                'totalVolume': total_v, 'totalOrders': total_o,
                'totalDefects': total_d, 'totalDefectOrders': total_do,
                'totalRate': round(total_d/total_v*100,2) if total_v>0 else 0,
                'totalOrderRate': round(total_do/total_o*100,2) if total_o>0 else 0,
                'factories': factory_rows_for_months(pm),
                'groupings': build_groupings_for_months(pm),
            }
    return {
        'key': key, 'name': display_name, 'label': period_label(mkeys), 'monthKeys': mkeys, 'months': labels,
        'monthlyVolume': vol, 'monthlyOrders': ords, 'monthlyDefects': defs, 'monthlyDefectOrders': def_orders, 'monthlyRate': rates,
        'totalVolume': total_vol, 'totalOrders': total_orders_p, 'totalDefects': total_defs, 'totalDefectOrders': total_def_orders,
        'totalRate': round(total_defs / total_vol * 100, 2) if total_vol > 0 else 0,
        'totalOrderRate': round(total_def_orders / total_orders_p * 100, 2) if total_orders_p > 0 else 0,
        'factories': rows,
        'groupings': build_groupings_for_months(mkeys),
        'prev': prev,
    }

PERIOD_DEFS = [
    ('all', 'All'),
    ('last_3', 'Last 3 months'),
    ('last_6', 'Last 6 months'),
    ('last_month', 'Last month'),
    ('mtd', 'MTD'),
    ('ytd', 'YTD'),
    ('quarter', 'Quarter'),
]
PERIODS = {k: build_period_payload(k, label) for k, label in PERIOD_DEFS}
PERIODS_JSON = json.dumps(PERIODS, cls=factory_data.DecimalEncoder)
PERIODS_JSON_SAFE = PERIODS_JSON.replace('<', '\\\\u003C').replace('>', '\\\\u003E')

# ── Remake Management data ──
remake_cur = conn.cursor()
remake_cur.execute("""
SELECT o.order_no,
       CAST(JSON_EXTRACT(o.price_info, '$.total_quantity') AS SIGNED) as qty,
       COALESCE(u.name, u.email, '(unknown)') AS admin_name,
       DATE_FORMAT(o.created_at, '%Y-%m') as month,
       COALESCE(GROUP_CONCAT(DISTINCT oi.factory_name ORDER BY oi.factory_name SEPARATOR ', '), '(unknown)') as factories
FROM orders o
LEFT JOIN users u ON u.id = o.order_administrator_id
LEFT JOIN order_items oi ON oi.order_id = o.id
WHERE o.order_type_symbol = 'R'
  AND o.created_at >= '2025-10-01'
  AND o.created_at < '2026-07-01'
GROUP BY o.order_no
ORDER BY qty DESC
""")
REMAKE_MGMT = [{"order": str(r[0]), "qty": int(r[1]) if r[1] else 0, "admin": r[2], "month": str(r[3])[:7] if r[3] else "?", "factory": r[4]} for r in remake_cur.fetchall()]
remake_cur.close()
REMAKE_MGMT_JSON = json.dumps(REMAKE_MGMT, cls=factory_data.DecimalEncoder)
conn.close()

# ── Remake Mgmt SAS token for universal save ──
from azure.storage.blob import BlobServiceClient, generate_blob_sas, BlobSasPermissions
from datetime import datetime, timedelta, timezone
AZURE_ACCOUNT = os.environ.get("AZURE_STORAGE_ACCOUNT", "custimoolivedata")
AZURE_KEY = os.environ.get("AZURE_STORAGE_KEY", "")
REMAKE_SAS_URL = ''
try:
    sas_token = generate_blob_sas(
        account_name=AZURE_ACCOUNT,
        container_name='$web',
        blob_name='remake-mgmt-data.json',
        account_key=AZURE_KEY,
        permission=BlobSasPermissions(read=True, write=True, create=True),
        expiry=datetime.now(timezone.utc) + timedelta(days=1),
        api_version='2021-12-02'
    )
    REMAKE_SAS_URL = f'https://{AZURE_ACCOUNT}.blob.core.windows.net/$web/remake-mgmt-data.json?{sas_token}'
except Exception as e:
    print("Warning: could not generate SAS token:", e)
    REMAKE_SAS_URL = ''



ORDERS_JSON = json.dumps(order_details, default=str)
ORDERS_JSON_SAFE = ORDERS_JSON.replace('<', '\\u003C').replace('>', '\\u003E')
FU_REVIEW_JSON = json.dumps(fu_review_rows, default=str)
FU_REVIEW_JSON_SAFE = FU_REVIEW_JSON.replace('<', '\\u003C').replace('>', '\\u003E')

MONTH_KEYS = json.dumps(all_months_sorted)
generation_time = datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')

html = f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Custimoo — Defect Report</title>
<script src="https://cdn.jsdelivr.net/npm/chart.js"></script>
<style>
  :root {{
    --bg: #f5f7fb;
    --card: #ffffff;
    --text: #172033;
    --muted: #667085;
    --border: #e6eaf2;
    --accent: #1f6feb;
    --shadow: 0 10px 30px rgba(16, 24, 40, 0.08);
    --radius: 18px;
  }}
  * {{ box-sizing: border-box; }}
  body {{ margin: 0; font-family: Inter, Segoe UI, Arial, sans-serif; background: var(--bg); color: var(--text); }}
  .wrap {{ max-width: 1100px; margin: 0 auto; padding: 24px; }}
  .hero {{ background: linear-gradient(135deg, #0f172a 0%, #1d4ed8 100%); color: #fff; border-radius: 24px; padding: 26px 28px; box-shadow: var(--shadow); margin-bottom: 18px; }}
  .hero h1 {{ margin: 0 0 6px; font-size: 28px; font-weight: 800; letter-spacing: -0.5px; }}
  .hero p {{ margin: 0; color: rgba(255,255,255,0.85); font-size: 13px; }}
  .tabs {{ display: flex; gap: 8px; margin-bottom: 18px; flex-wrap: wrap; }}
  .tab {{ background: #fff; color: var(--muted); border: 1px solid var(--border); border-radius: 999px; padding: 10px 18px; font-size: 13px; font-weight: 700; cursor: pointer; box-shadow: var(--shadow); transition: all 0.15s ease; }}
  .tab:hover {{ color: var(--accent); }}
  .tab.active {{ background: var(--accent); color: #fff; border-color: var(--accent); }}
  .page {{ display: none; }}
  .page.active {{ display: block; }}
  .exec-grid {{ display: grid; grid-template-columns: 1fr 1fr; gap: 16px; margin-bottom: 16px; }}
  .card {{ background: var(--card); border: 1px solid var(--border); border-radius: var(--radius); box-shadow: var(--shadow); padding: 22px; margin-bottom: 16px; }}
  .metric .label {{ color: var(--muted); font-size: 13px; margin-bottom: 10px; }}
  .metric .value {{ font-size: 48px; font-weight: 800; line-height: 1; margin-bottom: 8px; color: var(--accent); }}
  .metric .sub {{ color: var(--muted); font-size: 12px; line-height: 1.5; }}
  .section-title {{ font-size: 18px; font-weight: 800; margin: 0 0 14px; }}
  .section-head {{ display: flex; justify-content: space-between; align-items: center; margin-bottom: 14px; flex-wrap: wrap; gap: 8px; }}
  .section-head h3 {{ margin: 0; }}
  .reset-btn {{ background: #f3f4f6; color: var(--muted); border: 1px solid var(--border); border-radius: 8px; padding: 6px 12px; font-size: 12px; font-weight: 600; cursor: pointer; display: none; }}
  .reset-btn:hover {{ background: #e9ecf2; color: var(--text); }}
  .reset-btn.show {{ display: inline-block; }}
  .delta {{ font-size: 10px; margin-left: 4px; vertical-align: 1px; }}
  .delta.good {{ color: #16a34a; }}
  .delta.bad {{ color: #dc2626; }}
  .delta.neutral {{ color: #667085; }}
  table {{ width: 100%; border-collapse: collapse; font-size: 14px; }}
  th, td {{ padding: 11px 10px; border-bottom: 1px solid var(--border); text-align: left; }}
  th {{ font-size: 11px; text-transform: uppercase; letter-spacing: 0.04em; color: var(--muted); background: #fafbff; }}
  tr.clickable {{ cursor: pointer; transition: background 0.12s ease; }}
  tr.clickable:hover {{ background: #f0f6ff; }}
  tr.selected {{ background: #eaf2ff !important; box-shadow: inset 4px 0 0 var(--accent); }}
  .right {{ text-align: right; }}
  .pct-pill {{ display: inline-block; padding: 4px 12px; border-radius: 999px; font-weight: 700; font-size: 13px; cursor: pointer; transition: opacity 0.15s; }}
  .pct-pill:hover {{ opacity: 0.7; }}
  .pct-high {{ background: #fef3f2; color: #b42318; }}
  .pct-mid {{ background: #fffaeb; color: #b54708; }}
  .pct-low {{ background: #ecfdf3; color: #027a48; }}
  .chart-wrap {{ position: relative; height: 300px; width: 100%; }}
  .footnote {{ color: var(--muted); font-size: 12px; margin-top: 8px; font-style: italic; }}
  .trend-bar {{ display: flex; align-items: center; gap: 12px; margin-bottom: 12px; padding: 10px 14px; background: #f9fafb; border-radius: 10px; flex-wrap: wrap; }}
  .trend-label {{ font-size: 13px; color: var(--muted); }}
  .trend-pill {{ display: inline-flex; align-items: center; gap: 6px; padding: 5px 12px; border-radius: 999px; font-weight: 700; font-size: 13px; }}
  .trend-up {{ background: #fef3f2; color: #b42318; }}
  .trend-down {{ background: #ecfdf3; color: #027a48; }}
  .trend-flat {{ background: #f3f4f6; color: #475467; }}
  .factory-name {{ font-weight: 700; font-size: 16px; color: var(--text); }}
  .hint {{ font-size: 12px; color: var(--accent); margin-bottom: 10px; }}
  ul.clean {{ margin: 0; padding-left: 20px; line-height: 1.9; font-size: 14px; }}
  .in-progress {{ color: #b54708; font-weight: 600; }}
  
  /* Drill-down panel */
  .drill-overlay {{ display: none; position: fixed; inset: 0; background: rgba(0,0,0,0.4); z-index: 1000; align-items: flex-start; justify-content: center; padding-top: 40px; }}
  .drill-overlay.show {{ display: flex; }}
  .drill-panel {{ background: var(--card); border-radius: var(--radius); box-shadow: 0 20px 60px rgba(0,0,0,0.25); width: 90%; max-width: 780px; max-height: 85vh; overflow-y: auto; padding: 24px; }}
  .drill-title {{ font-size: 18px; font-weight: 800; margin-bottom: 14px; }}
  .drill-close {{ float: right; background: #f3f4f6; border: none; border-radius: 8px; padding: 6px 12px; font-size: 13px; font-weight: 600; cursor: pointer; }}
  .drill-close:hover {{ background: #e5e7eb; }}
  .drill-count {{ font-size: 13px; color: var(--muted); margin-bottom: 12px; }}
  .drill-table {{ font-size: 13px; }}
  .drill-table td {{ padding: 8px 8px; vertical-align: top; }}
  .drill-table .order-num {{ font-weight: 600; color: var(--accent); }}
  .drill-snippet {{ font-size: 11px; color: var(--muted); line-height: 1.4; max-height: 40px; overflow: hidden; }}
  @media (max-width: 700px) {{ .exec-grid {{ grid-template-columns: 1fr; }} .wrap {{ padding: 16px; }} .metric .value {{ font-size: 38px; }} }}
</style>
</head>
<body>
<div id="refresh-bar" style="text-align:center;padding:8px 16px;background:#1a1a2e;border-bottom:1px solid #2a2a4e;font-size:13px;color:#aaa;">
  <span id="last-update">Generated: {generation_time}</span>
  <button id="refresh-btn" onclick="doRefresh()" style="margin-left:12px;padding:4px 16px;background:#0f3460;color:white;border:1px solid #16213e;border-radius:4px;cursor:pointer;">Refresh Report</button>
  <a href="/dqc" style="margin-left:12px;color:#4ecca3;font-weight:700;text-decoration:none;">Open Digital QC Usage</a>
  <span id="refresh-msg" style="margin-left:10px;"></span>
</div>
<script>
async function doRefresh(){{var b=document.getElementById('refresh-btn'),m=document.getElementById('refresh-msg');b.disabled=!0;b.textContent='Refreshing...';m.textContent='';try{{var r=await fetch('/api/refresh'),d=await r.json();if(d.ok){{m.textContent='✓ '+d.message;m.style.color='#4ecca3';var t=0,c=setInterval(async()=>{{var s=await fetch('/api/status'),sd=await s.json();if(sd.conclusion==='success'){{clearInterval(c);location.reload();}}if(++t>60)clearInterval(c);}},2e3)}}else{{m.textContent='✗ '+(d.error||'Failed');m.style.color='#e94560'}}}}catch(e){{m.textContent='✗ Network error';m.style.color='#e94560'}}b.disabled=!1;b.textContent='Refresh Report'}}
</script>
<div class="wrap">
  <div class="hero">
    <h1>Custimoo — Defect Report</h1>
    <p>Reporting Period: {report_month_labels[0]} – {report_month_labels[-1]} ({report_month_labels[-1]} still in progress)</p>
  </div>
  <div class="tabs">
    <button class="tab active" data-target="summary">Summary</button>
    <button class="tab" data-target="ytd">YTD 2026</button>
    <button class="tab" data-target="details">Details</button>
    <button class="tab" data-target="methodology">Methodology</button>
    <button class="tab" data-target="error-tracking">Error Tracking</button>
    <button class="tab" data-target="remake-mgmt">Remake Mgmt</button>
    <button class="tab" data-target="dqc-usage">DQC Usage</button>
  </div>
  <section id="summary" class="page active">
    <div class="exec-grid">
      <div class="card metric"><div class="label">3-Month Rolling Error Rate</div><div class="value" id="rollingRate"></div><div class="sub" id="rollingSub"></div></div>
      <div class="card metric"><div class="label">Total Error Rate (Since Measurement Started)</div><div class="value" id="totalRate"></div><div class="sub" id="totalSub"></div></div>
    </div>
    <div class="card">
      <div class="section-head"><h3 class="section-title" id="breakdownTitle">Error Rate Breakdown — Factories</h3><div style="display:flex;align-items:center;gap:8px;margin:0"><label for="periodFilter" class="muted" style="font-size:13px;font-weight:700">Period:</label><select id="periodFilter" class="filter-select"><option value="all">All</option><option value="last_3">Last 3 months</option><option value="last_6">Last 6 months</option><option value="last_month">Last month</option><option value="mtd">MTD</option><option value="ytd">YTD</option><option value="quarter">Quarter</option></select><label for="measureFilter" class="muted" style="font-size:13px;font-weight:700">Measure:</label><select id="measureFilter" class="filter-select"><option value="qty">Qty</option><option value="orders">No of Orders</option></select><label for="breakdownFilter" class="muted" style="font-size:13px;font-weight:700">Filter:</label><select id="breakdownFilter" class="filter-select"><option value="all">All</option><option value="factory">Factories</option><option value="sku">SKU</option><option value="sport">Sports</option><option value="category">Category</option><option value="admin">Order Admin</option></select></div></div>
      <div class="hint" id="breakdownHint">Factory view shows FU customer feedback only. Physical QC measures are temporarily hidden while the data is being reviewed.</div>
      <table id="factoryTable"><thead><tr><th>Factory</th><th class="right">Total Order QTY</th><th class="right">FU Defects QTY</th><th class="right">FU ERR% (QTY)</th><th class="right">Remake Orders</th><th class="right">Remake QTY</th><th class="right">Remake / Total Order QTY</th></tr></thead><tbody id="factoryBody"></tbody></table>
    </div>
    <div class="card">
      <div class="section-head"><h3 class="section-title" id="trendTitle">Monthly Trend &mdash; All Factories</h3><button class="reset-btn" id="resetBtn">&larr; Show all factories</button></div>
      <div class="trend-bar" id="trendBar" style="display:none">
        <span class="factory-name" id="trendFactory"></span>
        <span class="trend-label">Trend (Oct &rarr; Apr):</span>
        <span class="trend-pill" id="trendPill"></span>
        <span class="trend-label" id="trendDelta"></span>
      </div>
      <div class="chart-wrap"><canvas id="trendChart"></canvas></div>
      <div class="footnote">Click any bar or line point to drill into the orders for that month · {report_month_labels[-1]} still in progress</div>
    </div>
  </section>
  <section id="ytd" class="page">
    <div class="exec-grid">
      <div class="card metric"><div class="label">YTD 2026 Total Order QTY</div><div class="value" id="ytdVolume"></div><div class="sub">Jan – Jun 2026*</div></div>
      <div class="card metric"><div class="label" id="periodKpiLabel">YTD 2026 Defects</div><div class="value" id="periodKpiValue"></div><div class="sub" id="periodKpiSub"></div></div>
    </div>
    <div class="card">
      <div class="section-head"><h3 class="section-title" id="ytdChartTitle">YTD Cumulative Total Order QTY</h3><div style="display:flex;align-items:center;gap:8px;margin:0"><label for="ytdMeasureFilter" class="muted" style="font-size:13px;font-weight:700">Measure:</label><select id="ytdMeasureFilter" class="filter-select"><option value="qty">Qty</option><option value="orders">No of Orders</option></select></div></div>
      <div class="chart-wrap"><canvas id="ytdChart"></canvas></div>
      <div class="footnote">Blue bars show accumulated volume/orders. Red line shows accumulated error percentage for the selected measure.</div>
    </div>
    <div class="card">
      <h3 class="section-title" id="ytdMonthlyTitle">YTD Monthly Accumulated QTY</h3>
      <table><thead><tr id="ytdMonthlyHead"></tr></thead><tbody id="ytdMonthlyBody"></tbody></table>
    </div>
    <div class="card">
      <h3 class="section-title">YTD 2026 &mdash; Per-Factory Error Rate (Worst &rarr; Best)</h3>
      <table><thead><tr id="ytdFactoryHead"></tr></thead><tbody id="ytdFactoryBody"></tbody></table>
    </div>
  </section>
  <section id="details" class="page">
    <div class="card">
      <h3 class="section-title">Month-wise Summary</h3>
      <div class="hint">Click any month row to see the defect orders</div>
      <table><thead><tr><th>Month</th><th class="right">Total Order QTY</th><th class="right">Orders</th><th class="right">Defect QTY</th><th class="right">Defect Orders</th><th class="right">Remake Orders</th><th class="right">Remake QTY</th><th class="right">Remake / Total Order QTY</th><th class="right">Err% (QTY)</th><th class="right">Err% (Orders)</th></tr></thead><tbody id="monthlyBody"></tbody></table>
    </div>
    <div class="card">
      <h3 class="section-title">Factory &times; Month Defects</h3>
      <div class="hint">Click any cell to see that factory's orders in that month</div>
      <table><thead><tr><th>Factory</th><th class="right" id="hdr1"></th><th class="right" id="hdr2"></th><th class="right" id="hdr3"></th><th class="right" id="hdr4"></th><th class="right" id="hdr5"></th><th class="right" id="hdr6"></th><th class="right" id="hdr7"></th><th class="right" id="hdr8"></th><th class="right" id="hdr9"></th></tr></thead><tbody id="factoryMonthBody"></tbody></table>
      <div class="footnote">{report_month_labels[-1]} still in progress</div>
    </div>
    <div class="card">
      <h3 class="section-title">FU Inbox Review</h3>
      <div class="hint">Orders read from fu@custimoo.com. Rows marked Counted affect the defect rate; delay/process or missing-quantity rows are tracked but not counted.</div>
      <table><thead><tr><th>FU Month</th><th>Order</th><th>Status</th><th class="right">Affected</th><th class="right">Total QTY</th><th>Factory</th><th>Subject</th></tr></thead><tbody id="fuReviewBody"></tbody></table>
    </div>
  </section>
  <section id="methodology" class="page">
    <div class="card">
      <h3 class="section-title">Included Product Groups</h3>
      <ul class="clean"><li>Jersey</li><li>Socks</li><li>Pants / Knickers</li><li>Shorts</li><li>Hoodie / Outerwear</li><li>Jacket</li><li>Shirt</li><li>Bags</li><li>Polo</li></ul>
    </div>
    <div class="card">
      <h3 class="section-title">How the Numbers Are Calculated</h3>
      <ul class="clean">
        <li>Reporting window is <strong>{report_month_labels[0]} – {report_month_labels[-1]}</strong>.</li>
        <li>Total month volume uses proper products only (excludes name plates, fight straps, logo patches, accessories).</li>
        <li>Defects are tracked from <strong>fu@custimoo.com</strong> customer feedback emails, with manual overrides for exact counts confirmed by order admins.</li>
        <li>When email does not specify the exact count, the full order quantity is used as the upper-bound estimate.</li>
        <li>Factory comparisons use total shipped order quantity per factory from the backend database, bucketed by <strong>order_items.status_updated_at</strong> for shipped/completed/shipping-status rows.</li>
        <li>Physical QC comparison measures are temporarily hidden while that data is being reviewed.</li>
        <li>Defects are bucketed by the month the customer <strong>first reported</strong> the issue (fu email received date), not the order creation date.</li>
        <li>{report_month_labels[-1]} is <span class="in-progress">still in progress</span>.</li>
        <li>Click any number in the report to drill into the specific orders behind it.</li>
        <li>Errors are tracked and estimated from the email <strong>fu@custimoo.com</strong>.</li>
        <li>We try to deduce the actual affected number of QTY to keep it fair. The system can sometimes have difficulties handling orders split across multiple factories.</li>
        <li>If an email is unclear or without order number, the sender will be asked to provide an order number and be a little more clear.</li>
      </ul>
    </div>
  </section>
  <section id="error-tracking" class="page">
    <div class="card">
      <div class="section-head"><h3 class="section-title">Error Tracking — Category Breakdown</h3><div style="display:flex;align-items:center;gap:8px;margin:0"><label for="errorFilter" class="muted" style="font-size:13px;font-weight:700">Group by:</label><select id="errorFilter" class="filter-select"><option value="factory">Factory</option><option value="sku">SKU</option><option value="sport">Sport</option><option value="admin">Order Admin</option></select></div></div>
      <div id="errorTrackingBody"></div>
    </div>
  </section>
  <section id="remake-mgmt" class="page">
    <div class="card">
      <h3 class="section-title">Remake Management — Order Admin Review</h3>
      <p class="muted">All remakes (no invoice) sorted by size. Categorize each to track reasons. Saved locally in your browser.</p>
      <div class="remake-filter-row" style="display:flex;gap:12px;align-items:center;margin-bottom:12px;flex-wrap:wrap">
        <label class="muted" style="font-size:13px;font-weight:700">Filter by admin:</label>
        <select id="remakeAdminFilter" class="filter-select" style="max-width:200px">
          <option value="">All admins</option>
        </select>
        <label class="muted" style="font-size:13px;font-weight:700">Filter by month:</label>
        <select id="remakeMonthFilter" class="filter-select" style="max-width:140px">
          <option value="">All months</option>
        </select>
        <span style="margin-left:auto;font-size:13px;color:var(--muted)" id="remakeCount">272 remakes</span>
        <button class="drill-close" id="remakeClearAll" style="font-size:12px;padding:4px 10px">Clear all saved</button>
      </div>
      <div style="overflow-x:auto;max-height:65vh;overflow-y:auto">
        <table class="remake-table"><thead>
          <tr><th>Order</th><th class="right">QTY</th><th>Admin</th><th>Factory</th><th>Month</th><th style="min-width:130px">Category</th><th style="min-width:120px">Fault</th><th style="min-width:200px">Comment</th><th></th></tr>
        </thead><tbody id="remakeMgmtBody"></tbody></table>
      </div>
    </div>
  </section>
  <section id="dqc-usage" class="page">
    <div class="card">
      <div class="section-head"><h3 class="section-title">Digital QC Usage</h3><div style="display:flex;align-items:center;gap:8px;margin:0;flex-wrap:wrap"><label class="muted" style="font-size:13px;font-weight:700">From:</label><input id="dqcFrom" type="date" class="filter-select" style="max-width:150px"><label class="muted" style="font-size:13px;font-weight:700">To:</label><input id="dqcTo" type="date" class="filter-select" style="max-width:150px"><button class="reset-btn" id="dqcRefreshBtn">Refresh</button><button class="reset-btn" id="dqcCsvBtn">CSV</button><button class="reset-btn" id="dqcXlsxBtn">Excel</button></div></div>
      <div class="hint" id="dqcGenerated">Loads audit runs from the central DQC logging API. Each row is one plugin audit run.</div>
    </div>
    <div class="exec-grid">
      <div class="card metric"><div class="label">Total Audits</div><div class="value" id="dqcTotal">–</div><div class="sub">Selected period</div></div>
      <div class="card metric"><div class="label">PASSED</div><div class="value" id="dqcPassed">–</div><div class="sub">Audit verdicts marked passed</div></div>
      <div class="card metric"><div class="label">REJECTED</div><div class="value" id="dqcRejected">–</div><div class="sub">Audit verdicts marked rejected</div></div>
      <div class="card metric"><div class="label">Users</div><div class="value" id="dqcUsers">–</div><div class="sub">Unique users running DQC</div></div>
    </div>
    <div class="card">
      <h3 class="section-title">Per-user Count</h3>
      <table><thead><tr><th>User</th><th class="right">Audits</th></tr></thead><tbody id="dqcUserBody"><tr><td colspan="2">Loading…</td></tr></tbody></table>
    </div>
    <div class="card">
      <h3 class="section-title">All DQC Runs</h3>
      <table><thead><tr><th>Date</th><th>User</th><th>Order</th><th>Verdict</th><th>Rejection Reason</th><th>DQC Skill Version</th><th>Timestamp UTC</th></tr></thead><tbody id="dqcRunBody"><tr><td colspan="7">Loading…</td></tr></tbody></table>
    </div>
  </section>

<!-- Drill-down overlay -->
<div class="drill-overlay" id="drillOverlay">
  <div class="drill-panel">
    <button class="drill-close" id="drillClose">✕ Close</button>
    <div class="drill-title" id="drillTitle">Orders</div>
    <div class="drill-count" id="drillCount"></div>
    <table class="drill-table" id="drillTable"><thead><tr><th>Order</th><th class="right">Affected</th><th>Product</th><th>Factory</th><th>Issue</th></tr></thead><tbody id="drillBody"></tbody></table>
  </div>
</div>

<script>
const DATA = {DATA_JSON};
const YTD = {YTD_DATA_JSON};
const FACTORY_COLORS = {FACTORY_COLORS};
const MONTH_KEYS = {MONTH_KEYS};
const ORDERS = {ORDERS_JSON_SAFE};
const FU_REVIEW = {FU_REVIEW_JSON_SAFE};
const GROUPINGS = {GROUPING_JSON_SAFE};
const ERROR_TRACKING = {ERROR_TRACKING_JSON_SAFE};
const PERIODS = {PERIODS_JSON_SAFE};
const REMAKES = {REMAKE_MGMT_JSON};
const REMAKE_SAVE_URL = '{REMAKE_SAS_URL}';
const REMAKE_DATA_URL = 'https://custimoolivedata.z13.web.core.windows.net/remake-mgmt-data.json';

const MONTH_LABELS = {{}};
MONTH_KEYS.forEach((k, i) => {{ MONTH_LABELS[k] = DATA.months[i]; }});
let ACTIVE_PERIOD = 'ytd';
let ACTIVE_DATA = PERIODS[ACTIVE_PERIOD] || DATA;
let ACTIVE_MONTH_KEYS = ACTIVE_DATA.monthKeys || MONTH_KEYS;
let ACTIVE_GROUPINGS = ACTIVE_DATA.groupings || GROUPINGS;
let ACTIVE_MEASURE = 'qty';
function activeMonthSet() {{ return new Set(ACTIVE_MONTH_KEYS); }}

// ── Utility ──
function ords() {{ return ORDERS; }}
function filterOrders(fn) {{ return ORDERS.filter(fn); }}

// ── Tabs ──
document.querySelectorAll('.tab[data-target]').forEach(function(btn) {{
  btn.addEventListener('click', function() {{
    document.querySelectorAll('.tab[data-target]').forEach(function(b) {{ b.classList.remove('active'); }});
    document.querySelectorAll('.page').forEach(function(p) {{ p.classList.remove('active'); }});
    btn.classList.add('active');
    document.getElementById(btn.dataset.target).classList.add('active');
    if (btn.dataset.target === 'ytd') setTimeout(renderYtdChart, 0);
    if (btn.dataset.target === 'dqc-usage') setTimeout(loadDqcUsage, 0);
    if (trendChart) setTimeout(function() {{ trendChart.resize(); }}, 0);
  }});
}});

// ── Drill-down ──
function dqcQs() {{
  var p = new URLSearchParams();
  var f = document.getElementById('dqcFrom');
  var t = document.getElementById('dqcTo');
  if (f && f.value) p.set('from', f.value);
  if (t && t.value) p.set('to', t.value);
  var s = p.toString();
  return s ? '?' + s : '';
}}
function dqcDownload(path) {{ window.location.href = path + dqcQs(); }}
async function loadDqcUsage() {{
  var msg = document.getElementById('dqcGenerated');
  if (!msg) return;
  msg.textContent = 'Loading DQC usage…';
  try {{
    var r = await fetch('/api/dqc/events' + dqcQs());
    var d = await r.json();
    if (!r.ok) throw new Error(d.error || r.statusText);
    var ev = d.events || [];
    var vc = {{PASSED:0, REJECTED:0, UNKNOWN:0}};
    var uc = {{}};
    ev.forEach(function(e) {{
      var v = (e.verdict || 'UNKNOWN').toUpperCase();
      vc[v] = (vc[v] || 0) + 1;
      var u = e.user || '(unknown)';
      uc[u] = (uc[u] || 0) + 1;
    }});
    document.getElementById('dqcTotal').textContent = ev.length.toLocaleString();
    document.getElementById('dqcPassed').textContent = (vc.PASSED || 0).toLocaleString();
    document.getElementById('dqcRejected').textContent = (vc.REJECTED || 0).toLocaleString();
    document.getElementById('dqcUsers').textContent = Object.keys(uc).length.toLocaleString();
    msg.textContent = 'API generated: ' + (d.generated_at || 'n/a') + ' · ' + ev.length.toLocaleString() + ' audit runs' + (d.stale_error ? ' · Warning: ' + d.stale_error : '');
    var users = Object.entries(uc).sort(function(a,b) {{ return b[1] - a[1]; }});
    document.getElementById('dqcUserBody').innerHTML = users.length ? users.map(function(x) {{ return '<tr><td>' + x[0] + '</td><td class="right">' + x[1].toLocaleString() + '</td></tr>'; }}).join('') : '<tr><td colspan="2">No users</td></tr>';
    document.getElementById('dqcRunBody').innerHTML = ev.length ? ev.map(function(e) {{
      var verdict = e.verdict || 'UNKNOWN';
      var reason = e.rejection_reason || e.reject_reason || e.reason || e.failure_reason || e.qc_reason || e.notes || e.message || '—';
      return '<tr><td>' + ((e.ts || '').slice(0,10)) + '</td><td>' + (e.user || '') + '</td><td>' + (e.order || '') + '</td><td><strong>' + verdict + '</strong></td><td>' + reason + '</td><td>0.5.5</td><td>' + (e.ts || '') + '</td></tr>';
    }}).join('') : '<tr><td colspan="7">No audits logged</td></tr>';
  }} catch(e) {{
    msg.innerHTML = '<span style="color:#b42318;font-weight:700">' + e.message + '</span>';
  }}
}}
var dqcRefreshBtn = document.getElementById('dqcRefreshBtn');
if (dqcRefreshBtn) dqcRefreshBtn.addEventListener('click', loadDqcUsage);
var dqcCsvBtn = document.getElementById('dqcCsvBtn');
if (dqcCsvBtn) dqcCsvBtn.addEventListener('click', function() {{ dqcDownload('/api/dqc.csv'); }});
var dqcXlsxBtn = document.getElementById('dqcXlsxBtn');
if (dqcXlsxBtn) dqcXlsxBtn.addEventListener('click', function() {{ dqcDownload('/api/dqc.xlsx'); }});

// ── Drill-down ──
function showDrill(title, orders) {{
  document.getElementById('drillTitle').textContent = title;
  document.getElementById('drillCount').textContent = orders.length + ' order' + (orders.length !== 1 ? 's' : '') + ' · ' + orders.reduce((s,o) => s + o.affected, 0).toLocaleString() + ' affected items';
  document.getElementById('drillBody').innerHTML = orders.map(o => 
    '<tr><td class="order-num">#' + o.order + '</td><td class="right">' + o.affected.toLocaleString() + '</td>'
    + '<td>' + o.product_type + '</td><td>' + o.factory + '</td>'
    + '<td style="font-size:12px;color:var(--muted);max-width:250px;white-space:nowrap;overflow:hidden;text-overflow:ellipsis" title="' + o.snippet.replace(/"/g,'&quot;') + '">' + o.subjects.slice(0, 60) + '</td></tr>'
  ).join('');
  document.getElementById('drillOverlay').classList.add('show');
}}

document.getElementById('drillClose').addEventListener('click', function() {{
  document.getElementById('drillOverlay').classList.remove('show');
}});
document.getElementById('drillOverlay').addEventListener('click', function(e) {{
  if (e.target.id === 'drillOverlay') document.getElementById('drillOverlay').classList.remove('show');
}});

function esc(s) {{ return String(s == null ? '' : s).replace(/[&<>"']/g, function(c) {{ return {{'&':'&amp;','<':'&lt;','>':'&gt;','"':'&quot;',"'":'&#39;'}}[c]; }}); }}
function renderFuReview() {{
  const body = document.getElementById('fuReviewBody');
  if (!body) return;
  const rows = (FU_REVIEW || []).slice(0, 250);
  body.innerHTML = rows.length ? rows.map(function(r) {{
    return '<tr><td>' + esc(r.fu_month) + '</td><td><strong>#' + esc(r.order) + '</strong></td><td>' + esc(r.status) + '</td><td class="right">' + esc(r.affected || '') + '</td><td class="right">' + (r.total_qty || 0).toLocaleString() + '</td><td>' + esc(r.factory) + '</td><td><div class="drill-snippet">' + esc(r.subjects || r.snippet || '') + '</div></td></tr>';
  }}).join('') : '<tr><td colspan="7">No FU messages read</td></tr>';
}}
renderFuReview();

// ── YTD KPI cards / measure view ──
function updateSummaryStats() {{
  const d = ACTIVE_DATA;
  const label = d.label || (d.months[0] + ' – ' + d.months[d.months.length-1]);
  document.getElementById('rollingRate').textContent = (d.totalRate || 0).toFixed(2) + '%';
  document.getElementById('rollingSub').textContent = (d.name || 'Selected period') + ' · ' + label;
  document.getElementById('totalRate').textContent = (d.totalRate || 0).toFixed(2) + '%';
  document.getElementById('totalSub').textContent = label + ' · ' + (d.totalDefects || 0).toLocaleString() + ' defects / ' + (d.totalVolume || 0).toLocaleString() + ' items';
}}


function pctPill(r) {{
  return r >= 2.0 ? '<span class="pct-pill pct-high">'+r.toFixed(2)+'%</span>'
       : r >= 1.0 ? '<span class="pct-pill pct-mid">'+r.toFixed(2)+'%</span>'
       : '<span class="pct-pill pct-low">'+r.toFixed(2)+'%</span>';
}}

function aggregateFactories(list) {{
  const totals = list.reduce(function(acc, f) {{
    const q = f.qarma || {{}};
    acc.volume += f.volume || 0;
    acc.orders += f.orders || 0;
    acc.defects += f.defects || 0;
    acc.defect_orders += f.defect_orders || 0;
    acc.remake_orders += f.remake_orders || 0;
    acc.remake_qty += f.remake_qty || 0;
    acc.qarma.sample_qty += q.sample_qty || 0;
    acc.qarma.defects += q.defects || 0;
    acc.qarma.inspections += q.inspections || 0;
    acc.qarma.orders_checked += q.orders_checked || 0;
    acc.qarma.rejected_orders += q.rejected_orders || 0;
    return acc;
  }}, {{volume:0, orders:0, defects:0, defect_orders:0, remake_orders:0, remake_qty:0, qarma:{{sample_qty:0, defects:0, inspections:0, orders_checked:0, rejected_orders:0}}}});
  totals.rate = totals.volume > 0 ? totals.defects / totals.volume * 100 : 0;
  totals.order_rate = totals.orders > 0 ? totals.defect_orders / totals.orders * 100 : 0;
  totals.qarma.rate = totals.qarma.sample_qty > 0 ? totals.qarma.defects / totals.qarma.sample_qty * 100 : 0;
  totals.qarma.order_rate = totals.qarma.orders_checked > 0 ? totals.qarma.rejected_orders / totals.qarma.orders_checked * 100 : 0;
  return totals;
}}


function compareMark(curr, prev, kind) {{
  if (prev === null || prev === undefined || isNaN(prev) || Number(prev) === 0 || curr === null || curr === undefined || isNaN(curr)) return '';
  const c = Number(curr), p = Number(prev);
  if (c === p) return '<span class="delta neutral" title="No change vs previous period">◆</span>';
  const up = c > p;
  let good = false;
  if (kind === 'bad') good = !up;
  else if (kind === 'good') good = up;
  else return '<span class="delta neutral" title="' + (up ? 'Up' : 'Down') + ' vs previous period">' + (up ? '▲' : '▼') + '</span>';
  return '<span class="delta ' + (good ? 'good' : 'bad') + '" title="' + (good ? 'Improved' : 'Worsened') + ' vs previous period">' + (up ? '▲' : '▼') + '</span>';
}}
function valWithDelta(html, curr, prev, kind) {{ return html; }}
function qarmaOrderRate(row) {{ const q = (row && row.qarma) || {{}}; return q.order_rate || 0; }}

function factoryRow(f, opts) {{
  opts = opts || {{}};
  const cls = opts.cls || '';
  const p = opts.prev || null;
  const clickable = opts.clickable ? ' clickable' : '';
  const dataFactory = opts.clickable ? ' data-factory="' + f.name + '"' : '';
  const remakeQtyPct = (f.volume || 0) > 0 ? (f.remake_qty || 0) / f.volume * 100 : 0;
  const pRemakeQtyPct = p && (p.volume || 0) > 0 ? (p.remake_qty || 0) / p.volume * 100 : null;
  const remakeOrderPct = (f.orders || 0) > 0 ? (f.remake_orders || 0) / f.orders * 100 : 0;
  const pRemakeOrderPct = p && (p.orders || 0) > 0 ? (p.remake_orders || 0) / p.orders * 100 : null;
  let row = '<tr class="' + (cls + clickable).trim() + '"' + dataFactory + '><td><strong>' + f.name + '</strong></td>';
  if (ACTIVE_MEASURE === 'orders') {{
    row += '<td class="right">' + valWithDelta((f.orders || 0).toLocaleString(), f.orders || 0, p ? p.orders || 0 : null, 'neutral') + '</td>'
      + '<td class="right">' + valWithDelta((f.defect_orders || 0).toLocaleString(), f.defect_orders || 0, p ? p.defect_orders || 0 : null, 'bad') + '</td>'
      + '<td class="right">' + valWithDelta(pctPill(f.order_rate || 0), f.order_rate || 0, p ? p.order_rate || 0 : null, 'bad') + '</td>'
      + '<td class="right">' + valWithDelta((f.remake_orders || 0).toLocaleString(), f.remake_orders || 0, p ? p.remake_orders || 0 : null, 'neutral') + '</td>'
      + '<td class="right">' + valWithDelta((f.remake_qty || 0).toLocaleString(), f.remake_qty || 0, p ? p.remake_qty || 0 : null, 'neutral') + '</td>'
      + '<td class="right">' + valWithDelta(pctPill(remakeOrderPct), remakeOrderPct, pRemakeOrderPct, 'bad') + '</td>';
  }} else {{
    row += '<td class="right">' + valWithDelta((f.volume || 0).toLocaleString(), f.volume || 0, p ? p.volume || 0 : null, 'neutral') + '</td>'
      + '<td class="right">' + valWithDelta((f.defects || 0).toLocaleString(), f.defects || 0, p ? p.defects || 0 : null, 'bad') + '</td>'
      + '<td class="right">' + valWithDelta(pctPill(f.rate || 0), f.rate || 0, p ? p.rate || 0 : null, 'bad') + '</td>'
      + '<td class="right">' + valWithDelta((f.remake_orders || 0).toLocaleString(), f.remake_orders || 0, p ? p.remake_orders || 0 : null, 'neutral') + '</td>'
      + '<td class="right">' + valWithDelta((f.remake_qty || 0).toLocaleString(), f.remake_qty || 0, p ? p.remake_qty || 0 : null, 'neutral') + '</td>'
      + '<td class="right">' + valWithDelta(pctPill(remakeQtyPct), remakeQtyPct, pRemakeQtyPct, 'bad') + '</td>';
  }}
  return row + '</tr>';
}}

function groupingRow(g, opts) {{
  opts = opts || {{}};
  const cls = opts.cls || '';
  const remakeOrderPct = (g.orders || 0) > 0 ? (g.remake_orders || 0) / g.orders * 100 : 0;
  const remakeQtyPct = (g.volume || 0) > 0 ? (g.remake_qty || 0) / g.volume * 100 : 0;
  let row = '<tr class="' + cls + '"><td><strong>' + g.name + '</strong></td>';
  if (ACTIVE_MEASURE === 'orders') {{
    row += '<td class="right">' + (g.orders || 0).toLocaleString() + '</td>'
      + '<td class="right">' + (g.defect_orders || 0).toLocaleString() + '</td>'
      + '<td class="right">' + pctPill(g.order_rate || 0) + '</td>'
      + '<td class="right">' + (g.remake_orders || 0).toLocaleString() + '</td>'
      + '<td class="right">' + (g.remake_qty || 0).toLocaleString() + '</td>'
      + '<td class="right">' + pctPill(remakeOrderPct) + '</td>';
  }} else {{
    row += '<td class="right">' + (g.volume || 0).toLocaleString() + '</td>'
      + '<td class="right">' + (g.defects || 0).toLocaleString() + '</td>'
      + '<td class="right">' + pctPill(g.rate || 0) + '</td>'
      + '<td class="right">' + (g.remake_orders || 0).toLocaleString() + '</td>'
      + '<td class="right">' + (g.remake_qty || 0).toLocaleString() + '</td>'
      + '<td class="right">' + pctPill(remakeQtyPct) + '</td>';
  }}
  return row + '</tr>';
}}

function setBreakdownHeader(mode) {{
  const thead = document.querySelector('#factoryTable thead tr');
  const first = mode === 'all' ? 'All' : (mode === 'factory' ? 'Factory' : (mode === 'sku' ? 'SKU / Series' : (mode === 'sport' ? 'Sport' : (mode === 'category' ? 'Category' : 'Order Admin'))));
  if (ACTIVE_MEASURE === 'orders') {{
    thead.innerHTML = '<th>' + first + '</th><th class="right">No of Orders</th><th class="right">FU Orders W/Defect</th><th class="right">FU ERR% (Orders)</th><th class="right">Remake Orders</th><th class="right">Remake QTY</th><th class="right">Remake %</th>';
  }} else {{
    thead.innerHTML = '<th>' + first + '</th><th class="right">Total Order QTY</th><th class="right">FU Defects QTY</th><th class="right">FU ERR% (QTY)</th><th class="right">Remake Orders</th><th class="right">Remake QTY</th><th class="right">Remake / Total Order QTY</th>';
  }}
  document.getElementById('breakdownTitle').textContent = mode === 'all' ? 'Error Rate Breakdown — All' : (mode === 'factory' ? 'Error Rate Breakdown — Factories' : (mode === 'sku' ? 'Error Rate Breakdown — SKU' : (mode === 'sport' ? 'Error Rate Breakdown — Sports' : (mode === 'category' ? 'Error Rate Breakdown — Category' : 'Error Rate Breakdown — Order Admin'))));
  document.getElementById('breakdownHint').textContent = mode === 'all' ? 'All selected period data in one total row.' : (mode === 'factory' ? 'Factory view shows FU customer feedback only. Physical QC measures are temporarily hidden while the data is being reviewed.' : (mode === 'category' ? 'Category view groups matched FU orders by high-confidence root-cause category; uncertain matches are Uncategorized. Physical QC measures are hidden for now.' : 'Same measure view as factory view; grouped by ' + (mode === 'sku' ? 'SKU / series' : (mode === 'sport' ? 'sport' : 'order admin')) + '. Physical QC measures are hidden for now.'));
}}

function renderFactoryTable(tbodyId, list, clickable, opts) {{
  const total = aggregateFactories(list);
  total.name = 'Total';
  const noMavicList = list.filter(function(f) {{ return f.name !== 'Mavic Sports'; }});
  const noMavic = aggregateFactories(noMavicList);
  noMavic.name = 'Total excl. Mavic Sports';
  const prevList = (opts && opts.prevList) || [];
  const prevByName = Object.fromEntries(prevList.map(function(x) {{ return [x.name, x]; }}));
  const prevTotal = prevList.length ? aggregateFactories(prevList) : null;
  const prevNoMavic = prevList.length ? aggregateFactories(prevList.filter(function(fd) {{ return fd.name !== 'Mavic Sports'; }})) : null;
  document.getElementById(tbodyId).innerHTML =
    list.map(function(f) {{ return factoryRow(f, {{clickable: clickable, prev: prevByName[f.name]}}); }}).join('')
    + factoryRow(total, {{cls:'total-row', prev: prevTotal}})
    + factoryRow(noMavic, {{cls:'no-mavic-row', prev: prevNoMavic}});
}}

function renderGroupingTable(mode) {{
  setBreakdownHeader(mode);
  var filter = document.getElementById('breakdownFilter'); if (filter && filter.value !== mode) filter.value = mode;
    if (mode === 'all') {{
    const total = aggregateFactories(ACTIVE_DATA.factories || []);
    total.name = 'All';
    let html = factoryRow(total, {{cls:'total-row'}});
    var prev = ACTIVE_DATA.prev;
    if (prev && prev.totalVolume > 0) {{
      const prevAgg = aggregateFactories(prev.factories || []);
      html += '<tr class="prev-row"><td><strong>Previous period (' + (prev.months || []).join(' \u2013 ') + ')</strong></td>';
      if (ACTIVE_MEASURE === 'orders') {{
        html += '<td class="right">' + (prev.totalOrders || 0).toLocaleString() + '</td>'
          + '<td class="right">' + (prev.totalDefectOrders || 0).toLocaleString() + '</td>'
          + '<td class="right">' + (prev.totalOrderRate || 0).toFixed(2) + '%</td>'
          + '<td class="right">' + (prevAgg.remake_orders || 0).toLocaleString() + '</td>'
          + '<td class="right">' + (prevAgg.remake_qty || 0).toLocaleString() + '</td>'
          + '<td class="right">' + pctPill((prevAgg.orders || 0) > 0 ? (prevAgg.remake_orders || 0) / prevAgg.orders * 100 : 0) + '</td>';
      }} else {{
        html += '<td class="right">' + (prev.totalVolume || 0).toLocaleString() + '</td>'
          + '<td class="right">' + (prev.totalDefects || 0).toLocaleString() + '</td>'
          + '<td class="right">' + (prev.totalRate || 0).toFixed(2) + '%</td>'
          + '<td class="right">' + (prevAgg.remake_orders || 0).toLocaleString() + '</td>'
          + '<td class="right">' + (prevAgg.remake_qty || 0).toLocaleString() + '</td>'
          + '<td class="right">' + pctPill((prevAgg.volume || 0) > 0 ? (prevAgg.remake_qty || 0) / prevAgg.volume * 100 : 0) + '</td>';
      }}
      html += '</tr>';
      var dVol = ACTIVE_DATA.totalVolume - prev.totalVolume;
      var dDef = ACTIVE_DATA.totalDefects - prev.totalDefects;
      html += '<tr class="delta-row"><td><strong>\u0394 vs prev</strong></td>';
      if (ACTIVE_MEASURE === 'orders') {{
        var dOrd = ACTIVE_DATA.totalOrders - prev.totalOrders;
        var dDefOrd = ACTIVE_DATA.totalDefectOrders - prev.totalDefectOrders;
        var dRemakeOrd = (total.remake_orders || 0) - (prevAgg.remake_orders || 0);
        var dRemakeQty = (total.remake_qty || 0) - (prevAgg.remake_qty || 0);
        html += '<td class="right">' + (dOrd >= 0 ? '+' : '') + dOrd.toLocaleString() + '</td>'
          + '<td class="right">' + (dDefOrd >= 0 ? '+' : '') + dDefOrd.toLocaleString() + '</td>'
          + '<td class="right">\u2014</td>'
          + '<td class="right">' + (dRemakeOrd >= 0 ? '+' : '') + dRemakeOrd.toLocaleString() + '</td>'
          + '<td class="right">' + (dRemakeQty >= 0 ? '+' : '') + dRemakeQty.toLocaleString() + '</td>'
          + '<td class="right">\u2014</td>';
      }} else {{
        var dRemakeOrd = (total.remake_orders || 0) - (prevAgg.remake_orders || 0);
        var dRemakeQty = (total.remake_qty || 0) - (prevAgg.remake_qty || 0);
        html += '<td class="right">' + (dVol >= 0 ? '+' : '') + dVol.toLocaleString() + '</td>'
          + '<td class="right">' + (dDef >= 0 ? '+' : '') + dDef.toLocaleString() + '</td>'
          + '<td class="right">\u2014</td>'
          + '<td class="right">' + (dRemakeOrd >= 0 ? '+' : '') + dRemakeOrd.toLocaleString() + '</td>'
          + '<td class="right">' + (dRemakeQty >= 0 ? '+' : '') + dRemakeQty.toLocaleString() + '</td>'
          + '<td class="right">\u2014</td>';
      }}
      html += '</tr>';
    }}
    document.getElementById('factoryBody').innerHTML = html;
    return;
  }}

  if (mode === 'factory') {{
    renderFactoryTable('factoryBody', ACTIVE_DATA.factories, true, {{prevList: (ACTIVE_DATA.prev && ACTIVE_DATA.prev.factories) || []}});
    return;
  }}
  const rows = ((ACTIVE_GROUPINGS || {{}})[mode] || []);
  const prevRows = (ACTIVE_DATA.prev && ACTIVE_DATA.prev.groupings && ACTIVE_DATA.prev.groupings[mode]) || [];
  const prevByName = Object.fromEntries(prevRows.map(function(x) {{ return [x.name, x]; }}));
  const total = aggregateFactories(rows);
  total.name = 'Total';
  const prevTotal = prevRows.length ? aggregateFactories(prevRows) : null;
  document.getElementById('factoryBody').innerHTML = rows.map(function(r) {{ return factoryRow(r, {{prev: prevByName[r.name]}}); }}).join('') + factoryRow(total, {{cls:'total-row', prev: prevTotal}});
}}

var breakdownFilter = document.getElementById('breakdownFilter');
if (breakdownFilter) {{
  breakdownFilter.addEventListener('change', function() {{
    var reset = document.getElementById('resetBtn');
    if (reset && reset.classList.contains('show')) reset.click();
    renderGroupingTable(breakdownFilter.value);
  }});
}}

function applyPeriod(key) {{
  ACTIVE_PERIOD = key;
  ACTIVE_DATA = PERIODS[key] || DATA;
  ACTIVE_MONTH_KEYS = ACTIVE_DATA.monthKeys || MONTH_KEYS;
  ACTIVE_GROUPINGS = ACTIVE_DATA.groupings || GROUPINGS;
  updateSummaryStats();
  updatePeriodKpis();
  var reset = document.getElementById('resetBtn');
  if (reset && reset.classList.contains('show')) reset.click();
  renderGroupingTable((document.getElementById('breakdownFilter') || {{value:'factory'}}).value);
  renderTrendChart(null);
}}
var periodFilter = document.getElementById('periodFilter');
if (periodFilter) {{
  periodFilter.value = ACTIVE_PERIOD;
  periodFilter.addEventListener('change', function() {{ applyPeriod(periodFilter.value); }});
}}

var measureFilter = document.getElementById('measureFilter');
if (measureFilter) {{
  measureFilter.value = ACTIVE_MEASURE;
  measureFilter.addEventListener('change', function() {{
    ACTIVE_MEASURE = measureFilter.value;
    renderGroupingTable((document.getElementById('breakdownFilter') || {{value:'factory'}}).value);
  }});
}}



// ── Factory tables ──
updateSummaryStats();
renderGroupingTable('factory');

// ── Monthly table (clickable rows for drill-down) ──
const mavicMonthly = DATA.factoryMonthly.find(function(fd) {{ return fd.name === 'Mavic Sports'; }}) || {{volumes:[], orders:[], defects:[], defect_orders:[], remake_orders:[], remake_qty:[]}};
function monthSummaryRow(label, qty, orders, defects, defectOrders, cls, remOrders, remQty) {{
  const qtyRate = qty > 0 ? defects / qty * 100 : 0;
  const orderRate = orders > 0 ? defectOrders / orders * 100 : 0;
  const remakeQtyRate = qty > 0 ? (remQty || 0) / qty * 100 : 0;
  return '<tr class="' + (cls || '') + '"><td><strong>' + label + '</strong></td>'
    + '<td class="right"><strong>' + qty.toLocaleString() + '</strong></td>'
    + '<td class="right"><strong>' + orders.toLocaleString() + '</strong></td>'
    + '<td class="right"><strong>' + defects.toLocaleString() + '</strong></td>'
    + '<td class="right"><strong>' + defectOrders.toLocaleString() + '</strong></td>'
    + '<td class="right"><strong>' + (remOrders || 0).toLocaleString() + '</strong></td>'
    + '<td class="right"><strong>' + (remQty || 0).toLocaleString() + '</strong></td>'
    + '<td class="right"><strong>' + remakeQtyRate.toFixed(2) + '%</strong></td>'
    + '<td class="right"><strong>' + qtyRate.toFixed(2) + '%</strong></td>'
    + '<td class="right"><strong>' + orderRate.toFixed(2) + '%</strong></td></tr>';
}}
const noMavicMonthlyQty = DATA.months.map(function(_, i) {{ return DATA.monthlyVolume[i] - (mavicMonthly.volumes[i] || 0); }});
const noMavicMonthlyOrders = DATA.months.map(function(_, i) {{ return DATA.monthlyOrders[i] - (mavicMonthly.orders[i] || 0); }});
const noMavicMonthlyDefects = DATA.months.map(function(_, i) {{ return DATA.monthlyDefects[i] - (mavicMonthly.defects[i] || 0); }});
const noMavicMonthlyDefectOrders = DATA.months.map(function(_, i) {{ return DATA.monthlyDefectOrders[i] - (mavicMonthly.defect_orders[i] || 0); }});
const noMavicTotals = {{
  qty: noMavicMonthlyQty.reduce(function(a,b) {{ return a+b; }}, 0),
  orders: noMavicMonthlyOrders.reduce(function(a,b) {{ return a+b; }}, 0),
  defects: noMavicMonthlyDefects.reduce(function(a,b) {{ return a+b; }}, 0),
  defectOrders: noMavicMonthlyDefectOrders.reduce(function(a,b) {{ return a+b; }}, 0)
}};
document.getElementById('monthlyBody').innerHTML = DATA.months.map((m, i) => {{
  var mkey = MONTH_KEYS[i];
  var ordRate = DATA.monthlyOrders[i] > 0 ? (DATA.monthlyDefectOrders[i] / DATA.monthlyOrders[i] * 100).toFixed(2) : '0.00';
  var remQty = DATA.monthlyRemakeQty ? DATA.monthlyRemakeQty[i] : 0;
  var remRate = DATA.monthlyVolume[i] > 0 ? (remQty / DATA.monthlyVolume[i] * 100).toFixed(2) : '0.00';
  return '<tr class="clickable" data-month="' + mkey + '"><td>' + m + '</td>'
    + '<td class="right">' + DATA.monthlyVolume[i].toLocaleString() + '</td>'
    + '<td class="right">' + DATA.monthlyOrders[i].toLocaleString() + '</td>'
    + '<td class="right">' + DATA.monthlyDefects[i].toLocaleString() + '</td>'
    + '<td class="right">' + DATA.monthlyDefectOrders[i].toLocaleString() + '</td>'
    + '<td class="right">' + (DATA.monthlyRemakeOrders ? DATA.monthlyRemakeOrders[i].toLocaleString() : '0') + '</td>'
    + '<td class="right">' + remQty.toLocaleString() + '</td>'
    + '<td class="right">' + remRate + '%</td>'
    + '<td class="right">' + DATA.monthlyRate[i].toFixed(2) + '%</td>'
    + '<td class="right">' + ordRate + '%</td></tr>';
}}).join('');

var totalRemakeOrders = DATA.monthlyRemakeOrders ? DATA.monthlyRemakeOrders.reduce(function(a,b){{return a+b;}}, 0) : 0;
var totalRemakeQty = DATA.monthlyRemakeQty ? DATA.monthlyRemakeQty.reduce(function(a,b){{return a+b;}}, 0) : 0;
var totalRemakeRate = DATA.totalVolume > 0 ? (totalRemakeQty / DATA.totalVolume * 100).toFixed(2) : '0.00';
var noMavicRemakeOrders = totalRemakeOrders - ((mavicMonthly.remake_orders || []).reduce(function(a,b){{return a+b;}}, 0));
var noMavicRemakeQty = totalRemakeQty - ((mavicMonthly.remake_qty || []).reduce(function(a,b){{return a+b;}}, 0));
document.getElementById('monthlyBody').innerHTML += '<tr><td><strong>Total</strong></td><td class="right"><strong>' + DATA.totalVolume.toLocaleString() + '</strong></td><td class="right"><strong>' + DATA.totalOrders.toLocaleString() + '</strong></td><td class="right"><strong>' + DATA.totalDefects.toLocaleString() + '</strong></td><td class="right"><strong>' + DATA.totalDefectOrders.toLocaleString() + '</strong></td><td class="right"><strong>' + totalRemakeOrders.toLocaleString() + '</strong></td><td class="right"><strong>' + totalRemakeQty.toLocaleString() + '</strong></td><td class="right"><strong>' + totalRemakeRate + '%</strong></td><td class="right"><strong>' + DATA.totalRate.toFixed(2) + '%</strong></td><td class="right"><strong>' + DATA.totalOrderRate.toFixed(2) + '%</strong></td></tr>'
   + monthSummaryRow('Total excl. Mavic Sports', noMavicTotals.qty, noMavicTotals.orders, noMavicTotals.defects, noMavicTotals.defectOrders, 'no-mavic-row', noMavicRemakeOrders, noMavicRemakeQty);

// ── Factory x Month header ──
DATA.months.forEach(function(m, i) {{
  var el = document.getElementById('hdr'+(i+1));
  if (el) el.textContent = m.replace('*','');
}});

// ── Factory x Month body (clickable cells with drill-down) ──
function sumFactoryMonths(list) {{
  return DATA.months.map(function(_, i) {{
    return list.reduce(function(sum, fd) {{ return sum + (fd.defects[i] || 0); }}, 0);
  }});
}}
function factoryMonthRow(name, values, cls, clickable) {{
  return '<tr class="' + (cls || '') + '"><td><strong>' + name + '</strong></td>'
    + values.map(function(d, i) {{
       if (clickable) return '<td class="right clickable drill-cell" data-month="' + MONTH_KEYS[i] + '" data-factory="' + name + '">' + (d || '0') + '</td>';
       return '<td class="right">' + (d || '0') + '</td>';
    }}).join('') + '</tr>';
}}
const totalFactoryMonth = sumFactoryMonths(DATA.factoryMonthly);
const noMavicFactoryMonth = sumFactoryMonths(DATA.factoryMonthly.filter(function(fd) {{ return fd.name !== 'Mavic Sports'; }}));
document.getElementById('factoryMonthBody').innerHTML = DATA.factoryMonthly.map(function(fd) {{
  return factoryMonthRow(fd.name, fd.defects, '', true);
}}).join('')
  + factoryMonthRow('Total', totalFactoryMonth, 'total-row', false)
  + factoryMonthRow('Total excl. Mavic Sports', noMavicFactoryMonth, 'no-mavic-row', false);


// ── YTD KPI cards / measure view ──
let YTD_MEASURE = 'qty';
document.getElementById('ytdVolume').textContent = DATA.totalVolume.toLocaleString();

function updatePeriodKpis() {{
  var d = ACTIVE_DATA;
  var label = d.name || 'YTD 2026';
  if (YTD_MEASURE === 'orders') {{
    document.getElementById('ytdVolume').textContent = (DATA.orders || 0).toLocaleString();
    document.getElementById('periodKpiValue').textContent = ((d.totalDefectOrders || 0)).toLocaleString();
    document.getElementById('periodKpiSub').textContent = label + ' · ' + (d.totalOrderRate || 0).toFixed(2) + '% of ' + (d.totalOrders || 0).toLocaleString() + ' orders · ' + (d.totalDefectOrders || 0).toLocaleString() + ' orders with defects';
  }} else {{
    document.getElementById('ytdVolume').textContent = (DATA.volume || 0).toLocaleString();
    document.getElementById('periodKpiValue').textContent = ((d.totalDefects || 0)).toLocaleString();
    document.getElementById('periodKpiSub').textContent = label + ' · ' + (d.totalRate || 0).toFixed(2) + '% of ' + (d.totalVolume || 0).toLocaleString() + ' items · ' + (d.totalDefectOrders || 0).toLocaleString() + ' orders with defects' + ' (' + (d.totalOrderRate || 0).toFixed(2) + '% orders)';
  }}
  document.getElementById('periodKpiLabel').textContent = label + ' Defects';
}}
function updateYtdKpis() {{
  updatePeriodKpis();
  if (YTD_MEASURE === 'orders') {{
    document.querySelector('#ytdVolume').closest('.metric').querySelector('.label').textContent = 'YTD 2026 No of Orders';
  }} else {{
    document.querySelector('#ytdVolume').closest('.metric').querySelector('.label').textContent = 'YTD 2026 Total Order QTY';
  }}
}}
updateYtdKpis();

// ── Charts ──
const chartBaseOptions = {{
  responsive: true,
  maintainAspectRatio: false,
  interaction: {{ mode: 'index', intersect: false }},
  plugins: {{ legend: {{ position: 'bottom' }} }},
  scales: {{
    y: {{ beginAtZero: true, title: {{ display: true, text: 'Total Order QTY' }} }},
    y1: {{ beginAtZero: true, position: 'right', grid: {{ drawOnChartArea: false }}, title: {{ display: true, text: 'Error %' }} }}
  }}
}};

let trendChart;
let currentTrendFactory = null;

function buildTrendDatasets(factoryName) {{
  if (!factoryName) {{
    const ds = [
      {{ type: 'bar', label: 'Total Order QTY', data: ACTIVE_DATA.monthlyVolume, backgroundColor: 'rgba(31, 111, 235, 0.25)', borderColor: 'rgba(31, 111, 235, 0.8)', borderWidth: 1, yAxisID: 'y' }},
      {{ type: 'line', label: 'Err% (Qty)', data: ACTIVE_DATA.monthlyRate, borderColor: '#ef4444', backgroundColor: '#ef4444', tension: 0.25, yAxisID: 'y1' }}
    ];
    return ds;
  }}
  const fd = (ACTIVE_DATA.factories || []).find(function(x) {{ return x.name === factoryName; }});
  if (!fd) return [];
  const vols = (fd.monthly && fd.monthly.volumes) || [];
  const defs = (fd.monthly && fd.monthly.defects) || [];
  const rates = vols.map(function(v, i) {{ return v > 0 ? +(defs[i] / v * 100).toFixed(2) : 0; }});
  return [
    {{ type: 'bar', label: factoryName + ' Total Order QTY', data: vols, backgroundColor: 'rgba(31, 111, 235, 0.25)', borderColor: 'rgba(31, 111, 235, 0.8)', borderWidth: 1, yAxisID: 'y' }},
    {{ type: 'line', label: factoryName + ' Err% (Qty)', data: rates, borderColor: FACTORY_COLORS[factoryName] || '#ef4444', backgroundColor: FACTORY_COLORS[factoryName] || '#ef4444', tension: 0.25, yAxisID: 'y1' }}
  ];
}}

function renderTrendChart(factoryName) {{
  currentTrendFactory = factoryName || null;
  document.getElementById('trendTitle').textContent = factoryName ? ('Monthly Trend — ' + factoryName) : 'Monthly Trend — All Factories';
  document.getElementById('resetBtn').style.display = factoryName ? 'inline-block' : 'none';
  document.getElementById('trendBar').style.display = factoryName ? 'flex' : 'none';
  if (factoryName) {{
    document.getElementById('trendFactory').textContent = factoryName;
    const fd = (ACTIVE_DATA.factories || []).find(function(x) {{ return x.name === factoryName; }});
    const vols2 = fd && fd.monthly ? fd.monthly.volumes : [];
    const defs2 = fd && fd.monthly ? fd.monthly.defects : [];
    const first = vols2[0] > 0 ? defs2[0] / vols2[0] * 100 : 0;
    const lastIdx = vols2.length - 1;
    const last = vols2[lastIdx] > 0 ? defs2[lastIdx] / vols2[lastIdx] * 100 : 0;
    const delta = last - first;
    document.getElementById('trendPill').textContent = last.toFixed(2) + '%';
    document.getElementById('trendDelta').textContent = (delta >= 0 ? '+' : '') + delta.toFixed(2) + ' pp vs first month';
  }}

  const ctx = document.getElementById('trendChart').getContext('2d');
  if (trendChart) trendChart.destroy();
  trendChart = new Chart(ctx, {{
    data: {{ labels: ACTIVE_DATA.months, datasets: buildTrendDatasets(factoryName) }},
    options: {{
      ...chartBaseOptions,
      onClick: function(evt, elements) {{
        if (!elements.length) return;
        const idx = elements[0].index;
        const mkey = ACTIVE_MONTH_KEYS[idx];
        const label = MONTH_LABELS[mkey] || mkey;
        const orders = filterOrders(function(o) {{ return o.fu_month === mkey && (!currentTrendFactory || o.factory === currentTrendFactory); }});
        showDrill((currentTrendFactory ? currentTrendFactory + ' — ' : '') + 'Orders reported in ' + label, orders);
      }}
    }}
  }});
}}
document.getElementById('resetBtn').addEventListener('click', function() {{ renderTrendChart(null); }});
document.querySelectorAll('#factoryBody tr[data-factory]').forEach(function(row) {{
  row.addEventListener('click', function() {{ renderTrendChart(row.dataset.factory); }});
}});

function ytdCumulativeTable() {{
  const head = document.getElementById('ytdMonthlyHead');
  const body = document.getElementById('ytdMonthlyBody');
  if (YTD_MEASURE === 'orders') {{
    head.innerHTML = '<th>Month</th><th class="right">Monthly No of Orders</th><th class="right">Accumulated No of Orders</th><th class="right">Monthly Defect Orders</th><th class="right">Accumulated Defect Orders</th><th class="right">Accumulated ERR% (Orders)</th>';
    body.innerHTML = YTD.months.map(function(m, i) {{
      return '<tr><td><strong>' + m + '</strong></td><td class="right">' + (YTD.monthlyOrders[i] || 0).toLocaleString() + '</td><td class="right">' + (YTD.cumulativeOrders[i] || 0).toLocaleString() + '</td><td class="right">' + (YTD.monthlyDefectOrders[i] || 0).toLocaleString() + '</td><td class="right">' + (YTD.cumulativeDefectOrders[i] || 0).toLocaleString() + '</td><td class="right">' + pctPill(YTD.cumulativeOrderRate[i] || 0) + '</td></tr>';
    }}).join('');
    document.getElementById('ytdMonthlyTitle').textContent = 'YTD Monthly Accumulated No of Orders';
  }} else {{
    head.innerHTML = '<th>Month</th><th class="right">Monthly Order QTY</th><th class="right">Accumulated Order QTY</th><th class="right">Monthly Defect QTY</th><th class="right">Accumulated Defect QTY</th><th class="right">Accumulated ERR% (QTY)</th>';
    body.innerHTML = YTD.months.map(function(m, i) {{
      return '<tr><td><strong>' + m + '</strong></td><td class="right">' + (YTD.monthlyVolume[i] || 0).toLocaleString() + '</td><td class="right">' + (YTD.cumulativeVolume[i] || 0).toLocaleString() + '</td><td class="right">' + (YTD.monthlyDefects[i] || 0).toLocaleString() + '</td><td class="right">' + (YTD.cumulativeDefects[i] || 0).toLocaleString() + '</td><td class="right">' + pctPill(YTD.cumulativeRate[i] || 0) + '</td></tr>';
    }}).join('');
    document.getElementById('ytdMonthlyTitle').textContent = 'YTD Monthly Accumulated QTY';
  }}
}}

function renderYtdFactoryTable() {{
  const head = document.getElementById('ytdFactoryHead');
  if (YTD_MEASURE === 'orders') {{
    head.innerHTML = '<th>Factory</th><th class="right">No of Orders</th><th class="right">FU Orders W/Defect</th><th class="right">FU ERR% (Orders)</th><th class="right">Remake Orders</th><th class="right">Remake QTY</th><th class="right">Remake %</th>';
  }} else {{
    head.innerHTML = '<th>Factory</th><th class="right">Total Order QTY</th><th class="right">FU Defects QTY</th><th class="right">FU ERR% (QTY)</th><th class="right">Remake Orders</th><th class="right">Remake QTY</th><th class="right">Remake / Total Order QTY</th>';
  }}
  const prev = ACTIVE_MEASURE;
  ACTIVE_MEASURE = YTD_MEASURE;
  renderFactoryTable('ytdFactoryBody', YTD.factories || [], false, {{}});
  ACTIVE_MEASURE = prev;
}}

let ytdChart = null;
function renderYtdChart() {{
  const el = document.getElementById('ytdChart');
  if (!el || !el.offsetParent) return;
  const isOrders = YTD_MEASURE === 'orders';
  const barData = isOrders ? YTD.cumulativeOrders : YTD.cumulativeVolume;
  const rateData = isOrders ? YTD.cumulativeOrderRate : YTD.cumulativeRate;
  const barLabel = isOrders ? 'Accumulated No of Orders' : 'Accumulated Total Order QTY';
  const rateLabel = isOrders ? 'Accumulated ERR% (Orders)' : 'Accumulated ERR% (QTY)';
  document.getElementById('ytdChartTitle').textContent = isOrders ? 'YTD Accumulated No of Orders + Error %' : 'YTD Accumulated Total Order QTY + Error %';
  if (ytdChart) ytdChart.destroy();
  ytdChart = new Chart(el.getContext('2d'), {{
    data: {{
      labels: YTD.months,
      datasets: [
        {{ type: 'bar', label: barLabel, data: barData, backgroundColor: 'rgba(31,111,235,0.25)', borderColor: 'rgba(31,111,235,0.8)', borderWidth: 1, yAxisID: 'y' }},
        {{ type: 'line', label: rateLabel, data: rateData, borderColor: '#ef4444', backgroundColor: '#ef4444', tension: 0.25, pointRadius: 5, pointHoverRadius: 7, yAxisID: 'y1' }}
      ]
    }},
    options: {{
      ...chartBaseOptions,
      scales: {{
        y: {{ beginAtZero: true, title: {{ display: true, text: barLabel }} }},
        y1: {{ beginAtZero: true, position: 'right', grid: {{ drawOnChartArea: false }}, title: {{ display: true, text: 'Accumulated Error %' }} }}
      }}
    }}
  }});
}}

function applyYtdMeasure() {{
  const sel = document.getElementById('ytdMeasureFilter');
  YTD_MEASURE = sel ? sel.value : 'qty';
  updateYtdKpis();
  ytdCumulativeTable();
  renderYtdFactoryTable();
  renderYtdChart();
}}
const ytdMeasureFilter = document.getElementById('ytdMeasureFilter');
if (ytdMeasureFilter) {{
  ytdMeasureFilter.value = YTD_MEASURE;
  ytdMeasureFilter.addEventListener('change', applyYtdMeasure);
}}
ytdCumulativeTable();
renderYtdFactoryTable();



// ── Drill click handlers ──
document.addEventListener('click', function(e) {{
  // Summary/YTD factory rows
  var factoryRow = e.target.closest('tr[data-factory]');
  if (factoryRow) {{
    var fname = factoryRow.dataset.factory;
    showDrill(fname + ' defect orders', filterOrders(function(o) {{ return o.factory === fname && activeMonthSet().has(o.fu_month); }}));
    return;
  }}

  // Monthly rows in Details tab
  var row = e.target.closest('tr[data-month]');
  if (row && !e.target.closest('.pct-pill')) {{
    var mkey = row.dataset.month;
    var label = MONTH_LABELS[mkey] || mkey;
    showDrill('Orders reported in ' + label, filterOrders(function(o) {{ return o.fu_month === mkey; }}));
    return;
  }}

  // Factory × Month cells
  var cell = e.target.closest('.drill-cell');
  if (cell) {{
    var cf = cell.dataset.factory;
    var cm = cell.dataset.month;
    var clabel = MONTH_LABELS[cm] || cm;
    showDrill(cf + ' defect orders in ' + clabel, filterOrders(function(o) {{ return o.factory === cf && o.fu_month === cm; }}));
    return;
  }}
}});

// ── Error Tracking tab ──
function renderErrorTracking(mode) {{
  const data = ERROR_TRACKING[mode] || {{}};
  const keys = Object.keys(data).sort();
  if (keys.length === 0) {{
    document.getElementById('errorTrackingBody').innerHTML = '<p class="muted">No error categories tracked for this group.</p>';
    return;
  }}
  let html = '';
  for (const group of keys) {{
    const cats = data[group];
    const totalQty = cats.reduce(function(acc, c) {{ return acc + (c.defect_qty || 0); }}, 0);
    html += '<h4 style="margin:20px 0 8px;font-size:15px;font-weight:700">' + group + ' <span class="muted" style="font-weight:400;font-size:12px">(' + cats.length + ' categories, ' + totalQty.toLocaleString() + ' total defects)</span></h4>';
    html += '<table class="cat-table"><thead><tr><th>Error Category</th><th class="right">Orders</th><th class="right">Defect QTY</th><th class="right">%</th><th class="right">Order #</th></tr></thead><tbody>';
    for (const cat of cats) {{
      const nums = cat.order_nums.slice(0,5).join(', ');
      const extra = cat.order_nums.length > 5 ? ' …+' + (cat.order_nums.length - 5) : '';
      const pct = totalQty > 0 ? (cat.defect_qty / totalQty * 100) : 0;
      html += '<tr><td><strong>' + cat.category + '</strong></td><td class="right">' + cat.order_count + '</td><td class="right">' + cat.defect_qty.toLocaleString() + '</td><td class="right">' + pct.toFixed(1) + '%</td><td class="right" style="font-size:11px;color:var(--muted);max-width:220px;overflow:hidden;text-overflow:ellipsis">#' + nums + extra + '</td></tr>';
    }}
    html += '</tbody></table>';
  }}
  document.getElementById('errorTrackingBody').innerHTML = html;
}}
// Init error tracking tab
const errorFilter = document.getElementById('errorFilter');
if (errorFilter) {{
  errorFilter.addEventListener('change', function() {{ renderErrorTracking(this.value); }});
  document.querySelectorAll('.tab[data-target="error-tracking"]').forEach(function(btn) {{
    btn.addEventListener('click', function() {{ setTimeout(function() {{ renderErrorTracking(errorFilter.value); }}, 0); }});
  }});
}}


// ── Remake Management ──
var remakeData = {{}};
var remakeSaveTimer = null;

function renderRemakeMgmt(filterAdmin, filterMonth) {{
  let rows = REMAKES;
  if (filterAdmin) rows = rows.filter(function(r) {{ return r.admin === filterAdmin; }});
  if (filterMonth) rows = rows.filter(function(r) {{ return r.month === filterMonth; }});
  const tbody = document.getElementById('remakeMgmtBody');
  tbody.innerHTML = rows.map(function(r) {{
    const key = r.order;
    const sd = remakeData[key] || {{category:'',fault:'',comment:''}};
    const catOpts = ['','Shipping','Quality','Delay','Bad design'].map(function(v) {{
      return '<option value="'+v+'"'+(sd.category===v?' selected':'')+'>'+(v||'\u2013')+'</option>';
    }}).join('');
    const faultOpts = ['','Custimoo','Customer','Other'].map(function(v) {{
      return '<option value="'+v+'"'+(sd.fault===v?' selected':'')+'>'+(v||'\u2013')+'</option>';
    }}).join('');
    return '<tr><td class="order-num">#'+r.order+'</td><td class="right">'+r.qty.toLocaleString()+'</td><td>'+r.admin+'</td><td>'+r.factory+'</td><td>'+r.month+'</td>'
      + '<td><select class="rm-cat" data-order="'+r.order+'" style="width:100%">'+catOpts+'</select></td>'
      + '<td><select class="rm-fault" data-order="'+r.order+'" style="width:100%">'+faultOpts+'</select></td>'
      + '<td><input class="rm-comment" data-order="'+r.order+'" type="text" value="'+(sd.comment||'')+'" style="width:100%;box-sizing:border-box" placeholder="Comment..."></td>'
      + '<td style="font-size:11px;color:#888">'+r.admin.split(' ')[0]+'</td></tr>';
  }}).join('');
  document.getElementById('remakeCount').textContent = rows.length+' remakes';
}}

function saveRemakeToAzure(order, key, val) {{
  if (!remakeData[order]) remakeData[order] = {{category:'',fault:'',comment:''}};
  remakeData[order][key] = val;
  // Debounce: wait 2s after last change then write to Azure
  if (remakeSaveTimer) clearTimeout(remakeSaveTimer);
  remakeSaveTimer = setTimeout(function() {{
    remakeSaveTimer = null;
    if (!REMAKE_SAVE_URL) return;
    fetch(REMAKE_SAVE_URL, {{
      method: 'PUT',
      headers: {{'Content-Type': 'application/json', 'x-ms-blob-type': 'BlockBlob'}},
      body: JSON.stringify(remakeData)
    }}).then(function(r) {{
      if (!r.ok) console.warn('Save failed:', r.status);
    }}).catch(function(e) {{
      console.warn('Save error:', e);
    }});
  }}, 2000);
}}

// Init Remake Mgmt
(function() {{
  if (!document.getElementById('remakeMgmtBody')) return;
  const admins = [...new Set(REMAKES.map(function(r){{return r.admin;}}))].sort();
  const months = [...new Set(REMAKES.map(function(r){{return r.month;}}))].sort();
  var af = document.getElementById('remakeAdminFilter');
  admins.forEach(function(a){{var opt=document.createElement('option');opt.value=a;opt.textContent=a;af.appendChild(opt);}});
  var mf = document.getElementById('remakeMonthFilter');
  months.forEach(function(m){{var opt=document.createElement('option');opt.value=m;opt.textContent=m;mf.appendChild(opt);}});
  // Load saved data from Azure
  fetch(REMAKE_DATA_URL + '?t=' + Date.now())
    .then(function(r) {{ if (r.ok) return r.json(); throw new Error('fetch failed'); }})
    .then(function(data) {{ remakeData = data || {{}}; }})
    .catch(function() {{ remakeData = {{}}; }})
    .finally(function() {{ renderRemakeMgmt('',''); }});
  af.addEventListener('change', function(){{renderRemakeMgmt(af.value, mf.value);}});
  mf.addEventListener('change', function(){{renderRemakeMgmt(af.value, mf.value);}});
  document.getElementById('remakeMgmtBody').addEventListener('change', function(e){{
    if (e.target.classList.contains('rm-cat')) saveRemakeToAzure(e.target.dataset.order,'category',e.target.value);
    if (e.target.classList.contains('rm-fault')) saveRemakeToAzure(e.target.dataset.order,'fault',e.target.value);
  }});
  document.getElementById('remakeMgmtBody').addEventListener('input', function(e){{
    if (e.target.classList.contains('rm-comment')) saveRemakeToAzure(e.target.dataset.order,'comment',e.target.value);
  }});
  document.getElementById('remakeClearAll').addEventListener('click', function(){{
    remakeData = {{}};
    if (REMAKE_SAVE_URL) {{
      fetch(REMAKE_SAVE_URL, {{
        method: 'PUT',
        headers: {{'Content-Type': 'application/json', 'x-ms-blob-type': 'BlockBlob'}},
        body: JSON.stringify({{}})
      }}).catch(function(e){{console.warn('Clear save error:', e);}});
    }}
    renderRemakeMgmt(af.value, mf.value);
  }});
  document.querySelectorAll('.tab[data-target="remake-mgmt"]').forEach(function(btn){{
    btn.addEventListener('click', function(){{setTimeout(function(){{renderRemakeMgmt(af.value, mf.value);}},0);}});
  }});
}})();
</script>
</body>
</html>"""

out_path = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), 'report.html')
with open(out_path, 'w') as f:
    f.write(html)
print("Written:", out_path, "(%d bytes)" % len(html))
