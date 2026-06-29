#!/usr/bin/env python3
import base64, csv, io, json, os, socketserver, urllib.parse, urllib.request
from collections import Counter
from datetime import datetime, timezone
from html import escape
import http.server

TOK = os.environ.get("GH_WORKFLOW_TOKEN", "")
REPO = "lars-lakr/custimoo-defect-report"
PORT = int(os.environ.get("PORT", 8080))
DQC_URL = os.environ.get("DQC_EVENTS_URL", "https://dqc-dashboard-custimoo.fly.dev/api/events")
DQC_USER = os.environ.get("DQC_DASH_USER", "")
DQC_PASS = os.environ.get("DQC_DASH_PASSWORD", "")
DQC_SKILL_VERSION = os.environ.get("DQC_SKILL_VERSION", "0.5.5")
REASON_KEYS = ("rejection_reason", "reject_reason", "reason", "failure_reason", "qc_reason", "notes", "message")

def event_reason(e):
    for k in REASON_KEYS:
        v = e.get(k)
        if v is not None and str(v).strip():
            return str(v).strip()
    return ""

USER_KEYS = ("windows_login", "windows_user", "windows_username", "login_name", "username", "user")

def event_user(e):
    for k in USER_KEYS:
        v = e.get(k)
        if v is not None and str(v).strip():
            return str(v).strip()
    return "(unknown)"

DQC_PAGE = """<!doctype html>
<html><head><meta charset="utf-8"><meta name="viewport" content="width=device-width, initial-scale=1">
<title>Custimoo DQC Usage</title>
<style>
body{font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',sans-serif;background:#f5f7fb;color:#172033;margin:0;padding:24px}.wrap{max-width:1200px;margin:auto}.top{display:flex;gap:12px;align-items:center;justify-content:space-between;flex-wrap:wrap}.card{background:#fff;border:1px solid #e6eaf2;border-radius:14px;padding:18px;margin:14px 0;box-shadow:0 8px 24px rgba(16,24,40,.06)}h1{margin:0 0 4px}.muted{color:#667085;font-size:13px}.filters{display:flex;gap:10px;flex-wrap:wrap;align-items:end}.filters label{font-size:12px;font-weight:700;color:#475467;display:block;margin-bottom:4px}input,button{padding:9px 10px;border:1px solid #d0d5dd;border-radius:8px;background:white}button{background:#0f3460;color:white;cursor:pointer;border-color:#0f3460}button.secondary{background:white;color:#0f3460}.grid{display:grid;grid-template-columns:repeat(4,minmax(140px,1fr));gap:12px}.kpi .label{font-size:12px;color:#667085;text-transform:uppercase;font-weight:700}.kpi .value{font-size:30px;font-weight:800;margin-top:4px}.passed{color:#07864b}.rejected{color:#c62828}table{width:100%;border-collapse:collapse;font-size:13px}th,td{padding:9px;border-bottom:1px solid #eef2f7;text-align:left}th{background:#f8fafc;font-weight:800}.right{text-align:right}.pill{display:inline-block;padding:3px 8px;border-radius:999px;font-weight:700;font-size:12px}.pill.PASSED{background:#e8f5ee;color:#087443}.pill.REJECTED{background:#fdeaea;color:#b42318}.error{color:#b42318;font-weight:700}@media(max-width:760px){.grid{grid-template-columns:1fr 1fr}}
</style></head><body><div class="wrap">
<div class="top"><div><h1>Custimoo Digital QC Usage</h1><div class="muted" id="generated">Loading…</div></div><div><a href="/" class="muted">← Failure report</a></div></div>
<div class="card filters"><div><label>From</label><input id="from" type="date"></div><div><label>To</label><input id="to" type="date"></div><div><button onclick="loadData()">Refresh</button></div><div><button class="secondary" onclick="download('/api/dqc.csv')">Download CSV</button></div><div><button class="secondary" onclick="download('/api/dqc.xlsx')">Download Excel</button></div></div>
<div id="msg" class="muted"></div>
<div class="grid"><div class="card kpi"><div class="label">Total audits</div><div class="value" id="total">–</div></div><div class="card kpi"><div class="label">PASSED</div><div class="value passed" id="passed">–</div></div><div class="card kpi"><div class="label">REJECTED</div><div class="value rejected" id="rejected">–</div></div><div class="card kpi"><div class="label">Users</div><div class="value" id="users">–</div></div></div>
<div class="card"><h3>Per-user count</h3><table><thead><tr><th>User</th><th class="right">Audits</th></tr></thead><tbody id="userBody"></tbody></table></div>
<div class="card"><h3>All runs</h3><table><thead><tr><th>Date</th><th>User</th><th>Order</th><th>Verdict</th><th>Rejection reason</th><th>DQC Skill Version</th><th>Timestamp UTC</th></tr></thead><tbody id="runBody"></tbody></table></div>
</div><script>
function qs(){const p=new URLSearchParams(); const f=document.getElementById('from').value,t=document.getElementById('to').value; if(f)p.set('from',f); if(t)p.set('to',t); return p.toString()?('?'+p.toString()):''}
function download(path){location.href=path+qs()}
async function loadData(){document.getElementById('msg').textContent='Loading…'; try{const r=await fetch('/api/dqc/events'+qs()); const d=await r.json(); if(!r.ok) throw new Error(d.error||r.statusText); render(d)}catch(e){document.getElementById('msg').innerHTML='<span class="error">'+e.message+'</span>'}}
function render(d){const ev=d.events||[]; document.getElementById('generated').textContent='API generated: '+(d.generated_at||'n/a')+' · '+ev.length+' rows'; document.getElementById('msg').textContent=d.stale_error?('Warning: '+d.stale_error):''; const vc={PASSED:0,REJECTED:0}; const uc={}; const reason=e=>(e.rejection_reason||e.reject_reason||e.reason||e.failure_reason||e.qc_reason||e.notes||e.message||'—'); const user=e=>(e.display_user||e.windows_login||e.windows_user||e.windows_username||e.login_name||e.username||e.user||'(unknown)'); ev.forEach(e=>{vc[(e.verdict||'').toUpperCase()]=(vc[(e.verdict||'').toUpperCase()]||0)+1; uc[user(e)]=(uc[user(e)]||0)+1}); document.getElementById('total').textContent=ev.length; document.getElementById('passed').textContent=vc.PASSED||0; document.getElementById('rejected').textContent=vc.REJECTED||0; document.getElementById('users').textContent=Object.keys(uc).length; document.getElementById('userBody').innerHTML=Object.entries(uc).sort((a,b)=>b[1]-a[1]).map(([u,c])=>`<tr><td>${u}</td><td class="right">${c}</td></tr>`).join('')||'<tr><td colspan=2>No users</td></tr>'; document.getElementById('runBody').innerHTML=ev.map(e=>`<tr><td>${(e.ts||'').slice(0,10)}</td><td>${user(e)}</td><td>${e.order||''}</td><td><span class="pill ${(e.verdict||'').toUpperCase()}">${e.verdict||''}</span></td><td>${reason(e)}</td><td>0.5.5</td><td>${e.ts||''}</td></tr>`).join('')||'<tr><td colspan=7>No audits logged</td></tr>'}
loadData();
</script></body></html>"""

class H(http.server.SimpleHTTPRequestHandler):
    def __init__(self, *args, **kwargs):
        super().__init__(*args, directory="/app", **kwargs)

    def do_GET(self):
        path = urllib.parse.urlparse(self.path).path
        if path == "/api/refresh": return self._refresh()
        if path == "/api/status": return self._status()
        if path == "/api/dqc/events": return self._dqc_events()
        if path == "/api/dqc.csv": return self._dqc_csv()
        if path == "/api/dqc.xlsx": return self._dqc_xlsx()
        if path == "/dqc": return self._html(DQC_PAGE)
        if path == "/":
            self.path = "/index.html"
            return super().do_GET()
        return super().do_GET()

    def _github_call(self, url, method="GET", body=None):
        headers = {"Authorization": "Bearer " + TOK, "Accept": "application/vnd.github+json"}
        if body: headers["Content-Type"] = "application/json"
        req = urllib.request.Request(url, data=body, method=method, headers=headers)
        with urllib.request.urlopen(req) as r:
            raw = r.read()
            return json.loads(raw) if raw else {}

    def _refresh(self):
        try:
            runs = self._github_call(f"https://api.github.com/repos/{REPO}/actions/runs?per_page=1&status=in_progress")
            if runs.get("workflow_runs"): return self._json(429, {"ok": False, "error": "Already running"})
            self._github_call(f"https://api.github.com/repos/{REPO}/actions/workflows/deploy.yml/dispatches", method="POST", body=json.dumps({"ref":"main"}).encode())
            return self._json(200, {"ok": True, "message": "Refresh triggered"})
        except Exception as e:
            return self._json(500, {"ok": False, "error": str(e)[:200]})

    def _status(self):
        try:
            data = self._github_call(f"https://api.github.com/repos/{REPO}/actions/runs?per_page=1&event=push&status=completed")
            runs = data.get("workflow_runs", [])
            return self._json(200, {"conclusion": runs[0]["conclusion"] if runs else "unknown", "updated_at": runs[0]["updated_at"] if runs else None})
        except Exception as e:
            return self._json(500, {"error": str(e)[:200]})

    def _dqc_query(self):
        incoming = urllib.parse.parse_qs(urllib.parse.urlparse(self.path).query)
        q = {"plugin": "custimoo-digital-qc"}
        for k in ("from", "to"):
            if incoming.get(k): q[k] = incoming[k][0]
        return urllib.parse.urlencode(q)

    def _fetch_dqc(self):
        if not DQC_USER or not DQC_PASS:
            raise RuntimeError("DQC credentials are not configured on the server")
        url = DQC_URL + "?" + self._dqc_query()
        token = base64.b64encode(f"{DQC_USER}:{DQC_PASS}".encode()).decode()
        req = urllib.request.Request(url, headers={"Authorization": "Basic " + token, "Accept": "application/json"})
        with urllib.request.urlopen(req, timeout=30) as r:
            data = json.loads(r.read().decode())
        events = data.get("events", []) or []
        for e in events:
            e["display_user"] = event_user(e)
        events.sort(key=lambda e: e.get("ts", ""), reverse=True)
        data["events"] = events
        data["summary"] = self._summarize(events)
        return data

    def _summarize(self, events):
        verdicts = Counter((e.get("verdict") or "UNKNOWN").upper() for e in events)
        users = Counter(event_user(e) for e in events)
        return {"total_audits": len(events), "verdicts": dict(verdicts), "users": dict(users)}

    def _dqc_events(self):
        try: return self._json(200, self._fetch_dqc())
        except Exception as e: return self._json(500, {"error": str(e)[:200]})

    def _dqc_csv(self):
        try:
            data = self._fetch_dqc(); out = io.StringIO(); w = csv.writer(out)
            w.writerow(["date", "user", "order", "verdict", "rejection_reason", "timestamp_utc", "dqc_skill_version"])
            for e in data.get("events", []): w.writerow([(e.get("ts") or "")[:10], event_user(e), e.get("order",""), e.get("verdict",""), event_reason(e), e.get("ts",""), DQC_SKILL_VERSION])
            return self._send(200, out.getvalue().encode(), "text/csv", "dqc_usage.csv")
        except Exception as e: return self._json(500, {"error": str(e)[:200]})

    def _dqc_xlsx(self):
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill
            data = self._fetch_dqc(); events = data.get("events", []); summary = data.get("summary", {})
            wb = Workbook(); ws = wb.active; ws.title = "Runs"
            headers = ["date", "user", "order", "verdict", "rejection_reason", "timestamp_utc", "dqc_skill_version"]
            ws.append(headers)
            for e in events: ws.append([(e.get("ts") or "")[:10], event_user(e), e.get("order",""), e.get("verdict",""), event_reason(e), e.get("ts",""), DQC_SKILL_VERSION])
            ws2 = wb.create_sheet("Summary"); ws2.append(["metric", "value"]); ws2.append(["total_audits", summary.get("total_audits", 0)]); ws2.append(["PASSED", summary.get("verdicts",{}).get("PASSED",0)]); ws2.append(["REJECTED", summary.get("verdicts",{}).get("REJECTED",0)]); ws2.append([]); ws2.append(["user", "audit_count"])
            for u,c in sorted(summary.get("users",{}).items(), key=lambda x: -x[1]): ws2.append([u,c])
            for sheet in (ws, ws2):
                for c in sheet[1]: c.font = Font(bold=True); c.fill = PatternFill("solid", fgColor="D9EAF7")
                for col in sheet.columns: sheet.column_dimensions[col[0].column_letter].width = min(max(len(str(cell.value or "")) for cell in col)+2, 45)
            bio = io.BytesIO(); wb.save(bio)
            return self._send(200, bio.getvalue(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", "dqc_usage.xlsx")
        except Exception as e: return self._json(500, {"error": str(e)[:200]})

    def _html(self, html): return self._send(200, html.encode(), "text/html; charset=utf-8")
    def _json(self, code, data): return self._send(code, json.dumps(data).encode(), "application/json")
    def _send(self, code, body, content_type, filename=None):
        self.send_response(code); self.send_header("Content-Type", content_type); self.send_header("Access-Control-Allow-Origin", "*"); self.send_header("Content-Length", str(len(body)))
        if filename: self.send_header("Content-Disposition", f'attachment; filename="{filename}"')
        self.end_headers(); self.wfile.write(body)
    def log_message(self, f, *a): pass

print(f"Server on :{PORT}", flush=True)
socketserver.TCPServer.allow_reuse_address = True
socketserver.TCPServer(("0.0.0.0", PORT), H).serve_forever()
